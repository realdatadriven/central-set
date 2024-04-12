'''ETL Module'''
# pylint: disable=unused-variable
# pylint: disable=line-too-long
# pylint: disable=wrong-import-order
# pylint: disable=broad-exception-caught
# pylint: disable=unused-import
# pylint: disable=invalid-name
# pylint: disable=trailing-whitespace
import os
import sys
import re
import copy
import tempfile
import json
import urllib
import zipfile
import datetime
import calendar
import pandas as pd
import numpy as np
import pyarrow.parquet as pq
from sqlalchemy import create_engine, Table, select, func, text
from sqlalchemy.engine.url import URL
from dateutil import parser
from ftplib import FTP
import openpyxl as xl
import sqlalchemy as sa
from query_doc import QueryDoc
import requests
from bs4 import BeautifulSoup
import duckdb
from fsspec import filesystem
import subprocess
import shutil
from pathlib import Path
from py_rust_odbc_csv import odbc_csv # pylint: disable=no-name-in-module
from werkzeug.utils import secure_filename
try:
    import win32com.client
except Exception as _err:
    win32com = None

from core.db import DB
from core.crud import Crud
from core.export import Export
from core.export_duckdb import ExportDucdb
from core.mail import Mail


class Etl:
    '''ETL Module'''
    def __init__(self, conf, params, db, i18n):
        self.conf = conf
        self.params = params
        self.db = db
        self.i18n = i18n
    # EXTRACT & LOAD | INPUTS
    async def extract(self):
        '''EXTRACT & LOAD | INPUTS'''
        try:
            _data = self.params['data']
            _input = _data['data']
            _etlrb = _data.get('selected_etlrb', {})
            _conf = {}
            _conf_etlrb = {}
            if not _input.get('active'):
                return {'success': False, 'msg': self.i18n('input-not-active', name = _input.get('etl_rbase_input', ''))}
            if _input.get('etl_rbase_input_conf'):
                try:
                    if isinstance(_input.get('etl_rbase_input_conf'), str):
                        _conf = json.loads(_input.get('etl_rbase_input_conf'))
                    elif isinstance(_input.get('etl_rbase_input_conf'), dict):
                        _conf = _input.get('etl_rbase_input_conf')
                    #print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            if _etlrb.get('etl_report_base_conf'):
                try:
                    if isinstance(_etlrb.get('etl_report_base_conf'), str):
                        _conf_etlrb = json.loads(_etlrb.get('etl_report_base_conf', {}))
                    elif isinstance(_etlrb.get('etl_report_base_conf'), dict):
                        _conf_etlrb = _etlrb.get('etl_report_base_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            fname = _input.get('file')
            is_tmp = _input.get('save_only_temp')
            date_ref = _input.get('date_ref')
            #print('date_ref:', date_ref, _conf)
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}'
            if is_tmp is True or _input.get('temp') or _input.get('tmp'):
                _path = tempfile.gettempdir()
            file_exists = None
            if fname:
                file_exists = os.path.exists(f'{_path}/{fname}')
            if fname and not file_exists:
                return  {'success': False, 'msg': self.i18n('file-not-founded', fname = fname)}
            if file_exists or _conf.get('type') in ['excel-parts', 'file-duckdb', 'file-to-duckdb', 'file-2-duckdb','duckdb']: # EXTRACT FROM FILE
                print('IS FILE', fname)
                if _conf.get('type') in ['file-duckdb', 'file-to-duckdb', 'file-2-duckdb']:
                    #"""```json
                    #{
                    #    "type": "file-duckdb",
                    #    "params": {
                    #        "database": "path 2 databse",
                    #        "extentions": ["spatial"],
                    #        "sql": "CREATE OR REPLACE TABLE TBL AS SELECT * FROM '<filename>'",
                    #    }
                    #}```"""
                    #print('FILE TO DUCKDB')
                    if not file_exists:
                        return {'success': False, 'msg': self.i18n('file-not-founded', fname = fname)}
                    if not _conf.get('duckdb'):
                        _conf['duckdb'] = copy.deepcopy(_conf.get('params'))
                    return await self._duckdb(_input, _etlrb, _conf, _conf_etlrb)
                if _conf.get('type') in ['duckdb'] or _conf.get('duckdb'):
                    if not _conf.get('duckdb'):
                        _conf['duckdb'] = copy.deepcopy(_conf.get('params'))
                    print('DUCKDB TO DUCKDB')
                    return await self._duckdb(_input, _etlrb, _conf, _conf_etlrb)
                elif _conf.get('type') == 'excel-parts':
                    #"""```json
                    #{
                    #    "type": "excel-parts",
                    #    "params": {
                    #        "sheet_name": "EoD Balances D",
                    #        "skiprows": 8, 
                    #        "nrows": null,
                    #        "usecols": [2, 10]
                    #    }
                    #}```"""
                    return await self._from_file(_input, _etlrb, _conf, _conf_etlrb)
                return await self._from_file(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'db': # EXTRACT FROM DB
                return await self._from_db(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'odbc-csv-duckdb': # EXTRACT FROM ODBC 2 CSV 2 DUCKDB
                return await self._from_odbc_csv_duckdb(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'scan_dir': # EXTRACT FROM A DIRECTORY
                return await self._from_ftp(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'ftp': # EXTRACT FROM FTP
                return await self._from_ftp(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'webscrap': # EXTRACT FROM WEBSCRAPING
                return await self._from_webscrap(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'outlook_mail': # OUTLOOK MAIL
                return await self._from_outlook(_input, _etlrb, _conf, _conf_etlrb)
            elif _conf.get('type') == 'list-ssform-duckdb': # LIST OS XLSX SSFORMS USING DUCKDB
                return await self._list_ssform_duckdb(_input, _etlrb, _conf, _conf_etlrb)
            else:
                if _conf.get('type'):
                    return {'success': False, 'msg': self.i18n('type-not-suported-yet', _type = _conf.get('type'))}
                else:
                    return {'success': False, 'msg': self.i18n('extract-failed')}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # TRANSFORM | OUTPUTS
    async def transform(self):
        '''TRANSFORM | OUTPUTS'''
        try:
            _data = self.params['data']
            _output = _data['data']
            if not _output.get('active'):
                return {'success': False, 'msg': self.i18n('output-not-active', name = _output.get('etl_rbase_output', ''))}
            _etlrb = _data.get('selected_etlrb', {})
            _step = _data.get('step', {})
            _conf = {}
            _conf_etlrb = {}
            if _output.get('etl_rbase_output_conf'):
                try:
                    if isinstance(_output.get('etl_rbase_output_conf'), str):
                        _conf = json.loads(_output.get('etl_rbase_output_conf', {}))
                    elif isinstance(_output.get('etl_rbase_output_conf'), dict):
                        _conf = _output.get('etl_rbase_output_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            if _conf_etlrb.get('etl_report_base_conf'):
                try:
                    if isinstance(_etlrb.get('etl_report_base_conf'), str):
                        _conf_etlrb = json.loads(_etlrb.get('etl_report_base_conf', {}))
                    elif isinstance(_etlrb.get('etl_report_base_conf'), dict):
                        _conf_etlrb = _etlrb.get('etl_report_base_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            date_ref = _step.get('dates_refs') if _step.get('dates_refs') else _output.get('date_ref')
            # print(date_ref, _step.get('dates_refs'), _output.get('date_ref'))
            # print(date_ref, _data.keys(), _output.keys(), _etlrb.keys(), _conf.keys())
            self.params['data']['table'] = _step.get('detail_table', 'etl_rb_output_field')
            self.params['data']['limit'] = -1
            self.params['data']['offset'] = 0
            self.params['data']['filters'] = [{
                'field': 'etl_rbase_output_id', 
                'cond': '=', 
                'value': _output.get('etl_rbase_output_id', 'none-founded')
            }]
            self.params['data']['order_by'] = [
                {'field': 'field_order',  'order': 'ASC'},
                {'field': 'etl_rb_output_field_id',  'order': 'ASC'}
            ]
            _crud = Crud(self.conf, self.params, self.db, self.i18n)
            _read = await _crud.read()
            if not _read.get('success'): # IN CASE OF ERR
                return  {'success': False, 'msg': self.i18n('output-fields-failed', name = _output.get('etl_rbase_output', ''), msg = _read.get('msg'))}
            elif len(_read.get('data', {})) == 0:                
                return {'success': False, 'msg': self.i18n('output-no-fields', name = _output.get('etl_rbase_output', ''), msg = _read.get('msg'))}
            else:
                _output_fields = _read.get('data', [])
                #print(list(map(lambda f: str(f.get('field_order')) + '->' + str(f.get('etl_rb_output_field')), _output_fields)))
                query_parts = {}
                for _field in _output_fields:
                    query_parts[_field['etl_rb_output_field']] = {
                        'name': _field['etl_rb_output_field'],
                        'desc': _field['etl_rb_output_field_desc'],
                        'select': _field['sql_select'],
                        'from_': _field['sql_from'],
                        'join': _field['sql_join'],
                        'where': _field['sql_where'],
                        'group_by': _field['sql_group_by'],
                        'order_by': _field['sql_order_by'],
                        'having': _field['sql_having'],
                        'window': _field['sql_window'],
                        'active': _field['active']
                    }
                _qd = QueryDoc(query_parts = query_parts)
                sql = _qd.get_query_sql(_parts = query_parts)
                if isinstance(date_ref, str):
                    date_ref = [ date_ref ]
                dates = [_dt if isinstance(_dt, (datetime.datetime, datetime.date)) else parser.parse(_dt) for _dt in date_ref]
                sql = _qd.set_date(sql, dates)
                if _output.get('output_type_id') == 1:                    
                    sql_drop = text(f'DROP TABLE IF EXISTS "{_output.get("destination_table")}"')
                    if _output.get('append_it', False) is True:
                        sql = text(f'CREATE IF NOT EXISTS TABLE "{_output.get("destination_table")}" AS {sql}')
                    else:
                        sql = text(f'CREATE OR REPLACE TABLE "{_output.get("destination_table")}" AS {sql}')
                elif _output.get('output_type_id') == 2:
                    sql_drop = text(f'DROP VIEW IF EXISTS "{_output.get("destination_table")}"')
                    sql = text(f'CREATE OR REPLACE VIEW "{_output.get("destination_table")}" AS {sql}')
                else:
                    return {'success': False, 'msg': self.i18n('output-type-not-suported', name = _output.get('etl_rbase_output', ''))}
                _database = copy.deepcopy(self.conf.get('DATABASE'))
                _database['database'] = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
                if _conf.get('params'):
                    _database = copy.deepcopy(_conf.get('params'))
                elif _conf_etlrb.get('params'):
                    _database = copy.deepcopy(_conf_etlrb.get('params'))
                _db = await self._get_new_db(_database)
                if isinstance(_db, sa.engine.base.Engine):
                    engine = _db
                else:
                    engine = _db.get_engine(_database)
                with engine.connect() as conn:
                    try:
                        if engine.driver in ('pysqlite', 'sqlite'):
                            cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                            conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                            busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000)# 60s / 1m
                            conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                        if _output.get('append_it', False) is True:
                            conn.execute(sql)
                            date_field_format = _output.get("date_field_format", 'YYYY-MM-DD')
                            sql = f'DELETE FROM "{_output.get("destination_table")}" WHERE "{_output.get("date_field")}" = \'{date_field_format}\''
                            sql = _qd.set_date(sql, dates)
                            #Sprint(sql)
                            conn.execute(text(sql))
                            sql = text(f'INSERT INTO "{_output.get("destination_table")}" VALUES {sql}')
                        conn.execute(sql)
                        conn.commit()
                        conn.close()
                        engine.dispose()
                        return {'success': True, 'msg': self.i18n('success')}
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF OUTPUT CREATION: ', str(_err), fname, exc_tb.tb_lineno)
                        return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err)), 'query_parts': query_parts}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # DATA QUALITY
    async def data_quality_check(self):
        '''data quality check'''
        try:
            return await self.data_quality({'only_check': True})
        except Exception as _err:# pylint: disable=broad-exception-caught
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    async def data_quality_fix(self):
        '''data quality fix'''
        try:
            return await self.data_quality({'only_fix': True})
        except Exception as _err:# pylint: disable=broad-exception-caught
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    async def data_quality(self, extras = None):
        '''data quality'''
        try:
            _data = self.params['data']
            _rule = _data.get('data', {})
            _etlrb = _data.get('selected_etlrb', {})
            _step = _data.get('step', {})
            _conf = {}
            _conf_etlrb = {}
            if _rule.get('etl_rbase_quality_conf'):
                try:
                    if isinstance(_rule.get('etl_rbase_quality_conf'), str):
                        _conf = json.loads(_rule.get('etl_rbase_quality_conf', {}))
                    elif isinstance(_rule.get('etl_rbase_quality_conf'), dict):
                        _conf = _rule.get('etl_rbase_quality_conf', {})
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            if _conf_etlrb.get('etl_report_base_conf'):
                try:
                    if isinstance(_etlrb.get('etl_report_base_conf'), str):
                        _conf_etlrb = json.loads(_etlrb.get('etl_report_base_conf', {}))
                    elif isinstance(_etlrb.get('etl_report_base_conf'), dict):
                        _conf_etlrb = _etlrb.get('etl_report_base_conf', {})
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            date_ref = _step.get('dates_refs') if _step.get('dates_refs') else _rule.get('date_ref')
            _data_quality_rules = []
            if _rule.get('etl_rbase_quality_id'):
                _data_quality_rules = [_rule]
            else:
                self.params['data']['table'] = _step.get('table', 'etl_rbase_quality')
                self.params['data']['limit'] = -1
                self.params['data']['offset'] = 0
                self.params['data']['filters'] = [{
                    'field': 'etl_report_base_id', 
                    'cond': '=', 
                    'value': _etlrb.get('etl_report_base_id', 'none-founded')
                }]
                _crud = Crud(self.conf, self.params, self.db, self.i18n)
                _read = await _crud.read()
                if not _read.get('success'):
                    return  {'success': False, 'msg': self.i18n('output-fields-failed', name = _rule.get('etl_rbase_output', ''), msg = _read.get('msg'))}
                elif len(_read.get('data', {})) == 0:             
                    return {'success': False, 'msg': self.i18n('output-no-fields', name = _rule.get('etl_rbase_output', ''), msg = _read.get('msg'))}
                else:
                    _data_quality_rules = _read.get('data', [])
            _qd = QueryDoc(query_parts = {})
            if isinstance(date_ref, str):
                date_ref = [ date_ref ]
            dates = [_dt if isinstance(_dt, (datetime.datetime, datetime.date)) else parser.parse(_dt) for _dt in date_ref]
            _cheks = {}
            _fixes = {}
            for _r in _data_quality_rules:
                if _r.get('active') and _r.get('sql_quality_check'):
                    sql = _r.get('sql_quality_check')
                    sql = _qd.set_date(sql, dates)
                    _cheks[_r['etl_rbase_quality_id']] = sql
                    if _r.get('sql_quality_fix'):
                        sql = _r.get('sql_quality_fix')
                        sql = _qd.set_date(sql, dates)
                        _fixes[_r['etl_rbase_quality_id']] = sql
            _extras = {}
            if extras:
                _extras = extras
            _database = copy.deepcopy(self.conf.get('DATABASE'))
            _database['database'] = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
            if _conf.get('params'):
                _database = copy.deepcopy(_conf.get('params'))
            elif _conf_etlrb.get('params'):
                _database = copy.deepcopy(_conf_etlrb.get('params'))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
            else:
                engine = _db.get_engine(_database)
            with engine.connect() as conn:
                try:
                    if engine.driver in ('pysqlite', 'sqlite'):
                        cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                        conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                        busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                        conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                    _data = {'check': {}, 'fix': {}}
                    patt_chk = re.compile(r'CREATE.*TABLE|UPDATE.*TABLE|DROP.*|INSERT.*INTO|DELETE|ALTER.*TABLE|UPSERT.*')
                    patt_fix = re.compile(r'CREATE.*TABLE|DROP.*|ALTER.*TABLE')
                    if _extras.get('only_check') is True or not extras:
                        for _q, _query in _cheks.items():
                            try:
                                _match = re.findall(patt_chk, _query)
                                if len(_match) > 0:
                                    _data['check'][_q] =  self.i18n('query-not-allowed', query = str(_query), match = '; '.join(_match))
                                    continue
                                res = conn.execute(text(_query))
                                _data['check'][_q] = res.mappings().all()[0].get('Total')
                                res.close()
                            except Exception as _err:# pylint: disable=broad-exception-caught
                                _data['check'][_q] = str(_err)
                    if _extras.get('only_fix') is True or not extras:
                        for _q, _query in _fixes.items():
                            try:
                                _match = re.findall(patt_fix, _query)
                                if len(_match) > 0:
                                    _data['fix'][_q] = self.i18n('query-not-allowed', query = str(_query), match = '; '.join(_match))
                                    continue
                                if _data['check'][_q] > 0:
                                    res = conn.execute(text(_query))
                                    _data['fix'][_q] = res.rowcount
                                    res.close()
                            except Exception as _err:# pylint: disable=broad-exception-caught
                                _data['fix'][_q] = str(_err)
                    conn.commit()
                    conn.close()
                    engine.dispose()
                    return {'success': True, 'msg': self.i18n('success'), 'data': _data}
                except Exception as _err:# pylint: disable=broad-exception-caught
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF OUTPUT CREATION: ', str(_err), fname, exc_tb.tb_lineno)
                    return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}  
        except Exception as _err:# pylint: disable=broad-exception-caught
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # RECONCILIATION 
    async def data_reconcilia(self):
        '''Reconciliation'''
        try:
            _data = self.params['data']
            _reconcil = _data['data']
            if not _reconcil.get('active'):
                return {'success': False, 'msg': self.i18n('reconc-not-active', name = _reconcil.get('etl_rb_reconcilia', ''))}
            _etlrb = _data.get('selected_etlrb', {})
            _step = _data.get('step', {})
            _conf = {}
            if _reconcil.get('etl_rb_reconcilia_conf'):
                try:
                    if isinstance(_reconcil.get('etl_rb_reconcilia_conf'), str):
                        _conf = json.loads(_reconcil.get('etl_rb_reconcilia_conf', {}))
                    elif isinstance(_reconcil.get('etl_rb_reconcilia_conf'), dict):
                        _conf = _reconcil.get('etl_rb_reconcilia_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            date_ref = _step.get('dates_refs') if _step.get('dates_refs') else _reconcil.get('date_ref')
            # print(date_ref, _data.keys(), _output.keys(), _etlrb.keys(), _conf.keys())
            self.params['data']['table'] = _step.get('detail_table', 'etl_rb_reconc_dtail')
            self.params['data']['limit'] = -1
            self.params['data']['offset'] = 0
            self.params['data']['filters'] = [{
                'field': 'etl_rb_reconcilia_id', 
                'cond': '=', 
                'value': _reconcil.get('etl_rb_reconcilia_id', 'none-founded')
            }]
            _crud = Crud(self.conf, self.params, self.db, self.i18n)
            _read = await _crud.read()
            if not _read.get('success'): # IN CASE OF ERR
                return  {'success': False, 'msg': self.i18n('reconc-detail-failed', name = _reconcil.get('etl_rb_reconcilia', ''), msg = _read.get('msg'))}
            elif len(_read.get('data', {})) == 0:                
                return {'success': False, 'msg': self.i18n('reconc-no-detail', name = _reconcil.get('etl_rb_reconcilia', ''), msg = _read.get('msg'))}
            else:
                _reconcil_details = _read.get('data', [])
                _parts = {}
                _qd = QueryDoc({})
                if isinstance(date_ref, str):
                    date_ref = [ date_ref ]
                dates = [_dt if isinstance(_dt, (datetime.datetime, datetime.date)) else parser.parse(_dt) for _dt in date_ref]               
                _database = copy.deepcopy(self.conf.get('DATABASE'))
                _database['database'] = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
                if _conf.get('params'):
                    _database = copy.deepcopy(_conf.get('params'))
                _db = await self._get_new_db(_database)
                if isinstance(_db, sa.engine.base.Engine):
                    engine = _db
                else:
                    engine = _db.get_engine(_database)
                with engine.connect() as conn:
                    try:
                        if engine.driver in ('pysqlite', 'sqlite'):
                            cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                            conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                            busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                            conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                        for _detail in _reconcil_details:
                            __id = _detail.get('etl_rb_reconc_dtail_id')
                            if not _detail.get('active'):
                                continue
                            if _detail.get('sql_query_val_1'):
                                sql = _qd.set_date(_detail.get('sql_query_val_1'), dates)
                                print(sql)
                                res = conn.execute(text(sql))
                                _detail['val_1'] = res.mappings().all()[0].get('Value')
                                res.close()                               
                            if _detail.get('sql_query_val_2'):
                                sql = _qd.set_date(_detail.get('sql_query_val_2'), dates)
                                print(sql)
                                res = conn.execute(text(sql))
                                _detail['val_2'] = res.mappings().all()[0].get('Value')
                                res.close()
                            if not _detail.get('is_eval_formula') and _detail.get('sql_reconcilia_query'):
                                sql = _qd.set_date(_detail.get('sql_reconcilia_query'), dates)
                                res = conn.execute(text(sql))
                                _detail['diff'] = res.mappings().all()[0].get('Value')
                                res.close()
                            else:
                                _detail['diff'] = _detail.get('sql_reconcilia_query')
                        conn.close()
                        engine.dispose()
                        if not self.params.get('data'):
                            self.params['data'] = {}
                        self.params['data']['data'] = _reconcil_details
                        self.params['data']['htmlstr'] = _reconcil.get('etl_rb_reconc_template')
                        _export = Export(self.conf, self.params, self.db, self.i18n)
                        _html = await _export.dump_2_html()
                        if not _html.get('success'):
                            return  {'success': False, 'msg': _html.get('msg')}
                        return {'success': True, 'msg': self.i18n('success'), 'html': _html.get('html'), 'data': _reconcil_details}
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF RECONC CREATION: ', str(_err), fname, exc_tb.tb_lineno)
                        return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err)), '_parts': _parts}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
        # Exports 
    async def export(self):
        '''Exports'''
        try:
            _data = self.params['data']
            _export = _data['data']
            if not _export.get('active'):
                return {'success': False, 'msg': self.i18n('export-not-active', name = _export.get('etl_rbase_export', ''))}
            _etlrb = _data.get('selected_etlrb', {})
            _step = _data.get('step', {})
            _conf = {}
            if _export.get('etl_rbase_export_conf'):
                try:
                    if isinstance(_export.get('etl_rbase_export_conf'), str):
                        _conf = json.loads(_export.get('etl_rbase_export_conf', {}))
                    elif isinstance(_export.get('etl_rbase_export_conf'), dict):
                        _conf = _export.get('etl_rbase_export_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            date_ref = _step.get('dates_refs') if _step.get('dates_refs') else _export.get('date_ref')
            if isinstance(date_ref, str):
                date_ref = [date_ref]
            #print(date_ref, _data.keys(), _output.keys(), _etlrb.keys(), _conf.keys())
            self.params['data']['table'] = _step.get('detail_table', 'etl_rb_exp_dtail')
            self.params['data']['limit'] = -1
            self.params['data']['offset'] = 0
            self.params['data']['filters'] = [{
                'field': 'etl_rbase_export_id', 
                'cond': '=', 
                'value': _export.get('etl_rbase_export_id', 'none-founded')
            }]
            _crud = Crud(self.conf, self.params, self.db, self.i18n)
            _read = await _crud.read()
            if not _read.get('success'): # IN CASE OF ERR
                return  {'success': False, 'msg': self.i18n('reconc-detail-failed', name = _export.get('etl_rb_exportia', ''), msg = _read.get('msg'))}
            elif len(_read.get('data', {})) == 0:                
                return {'success': False, 'msg': self.i18n('reconc-no-detail', name = _export.get('etl_rb_exportia', ''), msg = _read.get('msg'))}
            else:
                _export_details = _read.get('data', [])
                _qd = QueryDoc({})
                if isinstance(date_ref, str):
                    date_ref = [ date_ref ]
                dates = [_dt if isinstance(_dt, (datetime.datetime, datetime.date)) else parser.parse(_dt) for _dt in date_ref]               
                _exps = []
                for _dt in dates:
                    for _detail in _export_details:
                        #__id = _detail.get('etl_rb_exp_dtail_id')
                        if _detail.get('sql_export_query') and _detail.get('active'):
                            _detail['sql_export_query'] = _qd.set_date(_detail.get('sql_export_query'), _dt)
                            # print(_detail['sql_export_query'])
                    if not self.params.get('data'):
                        self.params['data'] = {}
                    self.params['data']['date_ref'] = _dt
                    self.params['data']['export'] = _export
                    self.params['data']['report'] = _etlrb
                    self.params['data']['_conf'] = _conf
                    self.params['data']['tmp'] = True
                    self.params['data']['data'] = _export_details
                    self.params['data']['file'] = _qd.set_date(_export.get('attach_file_template'), _dt) #_qd.set_date(_export.get('attach_file_template'), _dt)
                    self.params['data']['template'] = _export.get('attach_file_template')
                    self.params['database'] = _export.get('database', _etlrb.get('database'))
                    if _conf.get('duckdb_style') is True or len(re.findall(r'\.duckdb', self.params['database'])) > 0:
                        _exp_class = ExportDucdb(self.conf, self.params, self.db, self.i18n)
                        _exp = await _exp_class.export()
                    else:
                        _exp_class = Export(self.conf, self.params, self.db, self.i18n)
                        _exp = await _exp_class.export()        
                    _exp['date_ref'] = _dt.isoformat()
                    _exps.append(_exp)
                    _exps_has_err = list(filter(lambda _exp: _exp.get('success') is False, _exps))
                    if len(_exps_has_err) > 0:
                        return _exps_has_err[0]
                return {'success': True, 'msg': self.i18n('success'), 'data': _exps}   
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # Alert | Notefication
    async def notify(self):
        '''Alert | Notefication'''
        try:
            _qd = QueryDoc({})
            _data = self.params['data']
            _notify = _data['data']
            if not _notify.get('active'):
                return {'success': False, 'msg': self.i18n('notify-not-active', name = _notify.get('notify_subject', ''))}
            _etlrb = _data.get('selected_etlrb', {})
            _step = _data.get('step', {})
            _conf = {}
            _conf_etlrb = {}
            if _notify.get('notify_conf'):
                try:
                    if isinstance(_notify.get('notify_conf'), str):
                        _conf = json.loads(_notify.get('notify_conf', {}))
                    elif isinstance(_notify.get('notify_conf'), dict):
                        _conf = _notify.get('notify_conf', {})
                    # print('CONF:', _conf)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
            date_ref = _step.get('dates_refs') if _step.get('dates_refs') else _notify.get('date_ref')
            _params = _conf.get('params', {})
            _patt = re.compile(r'@ENV\..+')
            for _key in _params:
                match_env = re.findall(_patt, str(_params[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _params[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            _notes = []
            if isinstance(date_ref, str):
                date_ref = [ date_ref ]
            dates = [_dt if isinstance(_dt, (datetime.datetime, datetime.date)) else parser.parse(_dt) for _dt in date_ref] 
            if not self.params.get('data'):
                self.params['data'] = {}
            _notify['notify_copy_exports_path_bak'] = _notify.get('notify_copy_exports_path')
            logs = _data.get('db_data', {}).get('etl_report_base_log', {}).get('data', [])
            logs = pd.DataFrame(logs)
            logs['_ref'] = logs['ref']
            # print(logs.shape, logs.columns, logs.type.unique())
            if logs.shape[0] > 0:
                logs['_ref'] = np.vectorize(
                    lambda ref: (ref if isinstance(ref, (datetime.datetime, datetime.date)) else parser.parse(ref)).strftime('%Y%m%d')
                )(logs['ref'])
            for _dt in dates:
                _logs = logs.query(f"_ref == '{_dt.strftime('%Y%m%d')}'")
                if _notify.get('notify_copy_exports_path_bak'):
                    _notify['notify_copy_exports_path'] = _qd.set_date(_notify.get('notify_copy_exports_path_bak'), _dt)
                    _exported_files = []
                    if _notify.get('notify_copy_exports_path') and _notify.get('notify_copy_exports_to'):
                        # print(2, _notify.get('notify_copy_exports_path'), _notify.get('notify_copy_exports_to'), os.path.isdir(_notify.get('notify_copy_exports_path')))
                        if not os.path.isdir(_notify.get('notify_copy_exports_path')):
                            # print(_notify.get('notify_copy_exports_path'))
                            # os.makedirs(_notify.get('notify_copy_exports_path'))
                            Path(_notify.get('notify_copy_exports_path')).mkdir(parents = True, exist_ok = True)
                        _exports = _logs.query("type == 'EXPORT'")
                        for i,_exp in _exports.iterrows():
                            if _exp.get('fname'):
                                try:
                                    fname = _exp.get('fname')
                                    _, _filename = os.path.split(fname)
                                    _dest = f"{_notify.get('notify_copy_exports_path')}/{_filename}"
                                    _exported_files.append(fname)
                                    shutil.copy(f'{os.getcwd()}/{self.conf.get("UPLOAD")}/{fname}', _dest)
                                except Exception as _copy_err:
                                    print(_exp.get('fname'), str(_copy_err))
                    if _notify.get('notify_attach_exports'):
                        pass
                self.params['data']['data'] = {
                    'notify': _notify, 
                    'db_data': _data['db_data'], 
                    'logs': _logs.to_dict('records')
                }
                self.params['data']['htmlstr'] = _notify.get('notify_body')
                self.params['data']['_conf'] = _conf
                _export = Export(self.conf, self.params, self.db, self.i18n)
                _html = await _export.dump_2_html()
                if not _html.get('success'):
                    return  {'success': False, 'msg': _html.get('msg')}
                send_email = _notify.get('send_email') if 'send_email' in _notify.keys() else True
                if send_email is True:
                    #self.params['data']['mail']
                    self.params['mail'] = {
                        'smtp': _params.get('host'),
                        'port': _params.get('port'),
                        'from': _params.get('form', _params.get('username', _params.get('user'))),
                        'to': _notify.get('notify_to'),
                        'cc': _notify.get('notify_cc'),
                        'pass': _params.get('password', _params.get('pass')),
                        'subject': _qd.set_date(_notify.get('notify_subject'), _dt),
                        'type': 'html',
                        'body': _html.get('html'),
                        'attachments': _exported_files 
                            if _notify.get('notify_attach_exports') and len(_exported_files) > 0 
                            else None 
                    }
                    # print('mail:', self.params['mail'])
                    _mail = Mail(self.conf, self.params, self.db, self.i18n)
                    _smtp = await _mail.smtp_send()
                    if not _smtp.get('success'):
                        return _smtp
                    #_notes.append(self.params['data']['mail'])
                    _notes.append(self.params['mail'])
            return  {'success': True, 'msg': self.i18n('success'), 'data': _notes}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # RUN IMPORT
    async def _run_import(self, _df, _input, _etlrb, _conf, _conf_etlrb):
        '''Extract'''
        try:
            _total = _df.shape[0]
            n_cols = _df.shape[1]
            if _total == 0:
                return {'success': False, 'msg': self.i18n('no-rows-returned')}
            fname = _input.get('file', '')
            basename, ext = os.path.splitext(fname)
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            try:
                if _conf.get('column_names_2_upper') is True:
                    cols = [c.strip().upper() for c in _df.columns]
                else:
                    cols = [c.strip() for c in _df.columns]
                _df.columns = cols               
                _df.replace('None', None, regex = True, inplace = True)
            except Exception as _err:# pylint: disable=broad-exception-caught
                pass
            # FILE REF
            if 'file' not in _df and fname:
                _df['file'] = fname
            if fname:
                file_ref_patts = [
                    {'patt': r'(\d{8})(?!.*\d+)', 'fmrt': '%Y%m%d'},
                    {'patt': r'(\d{6})(?!.*\d+)', 'fmrt': '%Y%m'},
                    {'patt': r'(\d{4})(?!.*\d+)', 'fmrt': '%y%m'}
                ]
                for patt in file_ref_patts:
                    match = re.findall(patt.get('patt'), basename)
                    if len(match) > 0:
                        #print(str(match[0]))
                        try:
                            dt = datetime.datetime.strptime(str(match[0]), patt.get('fmrt'))
                            dt = datetime.date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
                            #print('file_ref:', patt.get('patt'), match[0], dt)
                            if 'file_ref' not in _df:
                                _df['file_ref'] = dt#.strftime('%Y-%m-%d')
                                _df['file_ref'] = _df['file_ref'].dt.date
                            break
                        except Exception as _err:# pylint: disable=broad-exception-caught
                            pass
            # TRANSFORM DATE FIELDS
            ref_date_field = _input.get('ref_date_field')
            other_date_fields = [] if not _input.get('other_date_fields', '') else _input.get('other_date_fields', '').split(',')
            if ref_date_field:
                other_date_fields.append(ref_date_field)
            if len(other_date_fields) > 0 and _input.get('date_format_org'):               
                for c in other_date_fields:
                    try:
                        if not c:
                            continue
                        elif c in _df:
                            dt_aux = _df[c].iloc[_df.index[0]]
                            print('other_date_fields:', c, dt_aux)
                            _already_right_dt_format = None
                            if isinstance(dt_aux, str):
                                _already_right_dt_format = re.match(r'\d{4}-\d{2}-\d{2}', dt_aux)
                            if not isinstance(dt_aux, (datetime.datetime, datetime.date)) and not _already_right_dt_format:
                                _df[c] = pd.to_datetime(_df[c], format = _input.get('date_format_org'), errors = 'coerce')
                                _df[c] = _df[c].dt.date
                            if isinstance(dt_aux, datetime.datetime):
                                _df[c] = _df[c].dt.date
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF: STR 2 DATA:', date_ref, str(_err), fname, exc_tb.tb_lineno)
                        return {'success': False, 'msg': self.i18n('unexpected-error', err = f'{str(c)}: {str(_err)}')}
            # DEST DATABASE
            _database = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
            if _conf_etlrb.get('params'):
                _database = copy.deepcopy(_conf_etlrb.get('params'))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
                metadata = sa.MetaData()
                #metadata.reflect(engine)
            else:
                engine = _db.get_engine(_database)
                metadata = self.db.get_metadata(engine)
            #print('DEST:', engine.url)
            # DATE VALIDATION
            check_ref_date = _input.get('check_ref_date')
            destination_table = _input.get('destination_table')
            sa_table = None
            try:
                sa_table = Table(destination_table, metadata, autoload_with = engine)
            except Exception as _err:# pylint: disable=broad-exception-caught
                pass
            if check_ref_date and ref_date_field and ref_date_field in _df:
                _df_date = _df[ref_date_field].unique()
                if len(_df_date) > 1:
                    return {
                        'success': False, 
                        'msg': self.i18n('date-valid-req-ref-field-1-date', 
                            dt_field = ref_date_field, 
                            dt_form = date_ref.strftime('%Y-%m-%d'), 
                            dt_file = ', '.join([parser.parse(d).strftime('%Y-%m-%d') if isinstance(d, str) else d.strftime('%Y-%m-%d') for d in _df_date])
                        )
                    }
                else:
                    print('_df_date:', _df_date)
                    _df_date = parser.parse(_df_date[0]).strftime('%Y-%m-%d') if not isinstance(_df_date[0], (datetime.date, datetime.datetime)) else _df_date[0].strftime('%Y-%m-%d')
                    if date_ref.strftime('%Y-%m-%d') != _df_date:
                        return {
                            'success': False, 
                            'msg': self.i18n('date-ref-form-diff-file', 
                                dt_field = ref_date_field, 
                                dt_form = date_ref.strftime('%Y-%m-%d'), 
                                dt_file = _df_date
                            )
                        }
                try:
                    if sa_table is not None:
                        sql = select(sa_table.c[ref_date_field]).\
                            select_from(sa_table).\
                            where(sa_table.c[ref_date_field] == date_ref.strftime('%Y-%m-%d')).\
                            limit(10)
                        with engine.connect() as conn:
                            if engine.driver in ('pysqlite', 'sqlite'):
                                cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                                conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                                busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                                conn.execute(text(f'PRAGMA busy_timeout = {str(busy_timeout)}'))
                            result = conn.execute(sql)
                            data = result.mappings().all()
                            result.close()
                            if len(data) > 0:
                                return {
                                    'success': False, 
                                    'msg': self.i18n('data-date-ref-exists',
                                        table = destination_table,
                                        dt_field = ref_date_field,
                                        dt_form = date_ref.strftime('%Y-%m-%d')
                                    )
                                }
                except Exception as _err:# pylint: disable=broad-exception-caught
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF: CHK DATA', date_ref, str(_err), fname, exc_tb.tb_lineno)
            # UPDATE DATABSE
            if_exists = 'replace' if _input.get('replace_existing_data', False) is True else 'append'
            _df['timestamp'] = datetime.datetime.now()
            with engine.begin() as conn:
                if engine.driver in ('pysqlite', 'sqlite'):
                    cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                    conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                    busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                    conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                if if_exists == 'append':
                    if sa_table is not None:
                        db_cols = [str(c).replace(f"{destination_table}.", "") for c in sa_table.columns]
                        if len(db_cols) > 0:
                            dtype = {
                                'int64': 'BIGINT',
                                'object': 'TEXT',
                                'float64': 'FLOAT',
                                'date': 'DATE',
                                'datetime64[ns]': 'DATETIME',
                                'datetime64': 'DATETIME',
                                'bool': 'BOOLEAN'
                            }                              
                            for c in _df.columns:
                                #print(c)
                                if c.find('Unnamed') != -1 and _conf.get('type') != 'excel-parts':
                                    _df.drop(c, axis = 'columns', inplace = True)
                                elif not c in db_cols:
                                    try:
                                        conn.execute(text(f'ALTER TABLE "{destination_table}" ADD COLUMN "{c}" {dtype[str(_df[c].dtype)]}'))
                                    except Exception as _err:# pylint: disable=broad-exception-caught
                                        *_, exc_tb = sys.exc_info()
                                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                        print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
                if if_exists == 'replace':         
                    try:
                        if engine.driver in ('postgres', 'pg'):
                            conn.execute(text(f"""DROP TABLE IF EXISTS "{destination_table}";"""))
                    except Exception as _err2:
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF [DROP BEFORE INSERT]: ', destination_table, str(_err2), fname, exc_tb.tb_lineno)               
                print('IMPORTING...', _df.shape)
                # _df.to_csv(f'{destination_table}.csv', index = False, sep = ';', decimal = ',')
                _df.to_sql(destination_table, con = conn, if_exists = if_exists, index = False, chunksize = 10000)
                conn.commit()
                conn.close()
            return {'success': True, 'msg': self.i18n('success'), 'n_rows': _total, 'n_cols': n_cols}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # EXTRACT FROM FILE
    async def _from_file(self, _input, _etlrb, _conf, _conf_etlrb):
        '''Import data from file'''
        try:        
            fname = _input.get('file')
            is_tmp = _input.get('save_only_temp')
            date_ref = _input.get('date_ref')
            columns_to_import = _input.get('columns_to_import', '').split(',') if _input.get('columns_to_import') else []
            if len(columns_to_import) < 2:
                columns_to_import = None
            engine = self.db.get_engine()
            headers = _input.get('headers', '').split(',') if _input.get('headers') else []
            _qd = QueryDoc(query_parts = {})
            if len(headers) < 2:
                headers = None
            else:
                is_ref = re.findall(r'^.+\[.+\]$', _input.get('headers', ''))
                if len(is_ref) > 0:
                    tbl = re.sub(r'\[.+\]$', '', _input.get('headers', ''))
                    fld = re.sub(r'\[|\]', '', _input.get('headers', ''))
                    print(_input.get('headers', ''), tbl, fld)
                    sql = text(f"""SELECT *\nFROM "{tbl}"\n""")
                    df_layout = pd.read_sql(sql , con = engine.connect())
                    headers = df_layout[fld].values
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}'
            if is_tmp is True or _input.get('temp') or _input.get('tmp'):
                _path = tempfile.gettempdir()
            df = pd.DataFrame([])
            basename, ext = os.path.splitext(fname)
            ext2 = ''
            if ext.lower() in ['.zip', '.gz', '.bz']:
                basename2, ext2 = os.path.splitext(basename)
            if _input.get('spreadsheet_forms') and _input.get('spreadsheet_forms_map'):
                _map = _input.get('spreadsheet_forms_map')
                zf = None
                z_files = []
                if ext.lower() in ['.zip']:
                    zf = zipfile.ZipFile(f'{_path}/{fname}')
                    z_files = zf.namelist()
                else:
                    _zf = zipfile.ZipFile(f'{_path}/{basename}.zip', 'w', zipfile.ZIP_DEFLATED)
                    _zf.write(f'{_path}/{fname}', fname)
                    _zf.close()
                    zf = zipfile.ZipFile(f'{_path}/{basename}.zip')
                    z_files = zf.namelist()
                mappin = {}
                map_has_versioning = False
                if _conf.get('type') == 'mapping' and _conf.get('map') and _conf.get('version'):
                    patt = re.compile(r'{.*}', re.IGNORECASE)
                    matches = re.findall(patt, _map)
                    matches = re.findall(patt, _conf.get('map'))
                    if len(matches) > 0:
                        map_has_versioning = True
                    elif len(matches) > 0:
                        map_has_versioning = True
                multi_maps = str(_map).split(',')
                if map_has_versioning is False:
                    try:
                        if len(multi_maps) > 1:
                            mappin = {}
                            for m in multi_maps:                       
                                sql = text(f'SELECT * FROM "{str(m).strip()}"')
                                mappin[str(m).strip()] = pd.read_sql(sql, engine)
                        else:
                            sql = text(f'SELECT * FROM "{_map.strip()}"')
                            mappin[str(_map).strip()] = pd.read_sql(sql, engine)
                    except Exception as e:
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF: ', str(e), fname, exc_tb.tb_lineno)
                proccessed_data = []
                if zf and len(z_files) > 0:
                    try:
                        for file in z_files:
                            basename, extension = os.path.splitext(file)
                            file_pth = zf.open(file, mode = 'r')
                            wb = xl.load_workbook(file_pth, read_only = True, keep_links = False, data_only = True)
                            version = ''
                            map_version = ''
                            new_map = ''
                            sheet = ''
                            is_static = False
                            if _conf.get('version') and isinstance(_conf.get('version'), dict):
                                sheet = _conf["version"].get('sheet')
                            elif _conf.get('version') and isinstance(_conf.get('version'), list):
                                is_static = True
                                for cnf in _conf.get('version'):
                                    try:
                                        if cnf.get('sheet') in wb.sheetnames:
                                            ws = wb[cnf.get('sheet')]
                                            if ws[cnf.get('range')].value == cnf.get('value'):
                                                version = str(cnf.get('version'))
                                                new_map = _conf.get("map").format(VERSION = str(version).upper())
                                                sheet = cnf.get('sheet')
                                                break
                                    except Exception as __err:
                                        pass
                            if map_has_versioning is True and (not sheet in wb.sheetnames):
                                patt = re.compile(r'_{.*}', re.IGNORECASE)
                                new_map = re.sub(patt, '', _conf["map"])
                                map_version = new_map
                                if not new_map in mappin:
                                    try:
                                        sql = text(f"""SELECT * FROM "{map_version}"\n""")
                                        mappin[map_version] = pd.read_sql(sql, engine)
                                    except Exception as e:
                                        *_, exc_tb = sys.exc_info()
                                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                        print('DEBUG INF: ', str(e), fname, exc_tb.tb_lineno)
                            elif map_has_versioning is True and sheet in wb.sheetnames:
                                if is_static is False:
                                    ws = wb[_conf["version"].get('sheet')]
                                    rng = ws[_conf["version"].get('range')]
                                    data = []
                                    for r in rng:
                                        _aux = []
                                        _all_empty = True
                                        for cell in r:
                                            _aux.append(cell.value)
                                            if cell.value:
                                                _all_empty = False
                                        if _all_empty is False:
                                            data.append(copy.deepcopy(_aux))
                                    cols = [c.strip() for c in data[0]]
                                    data.pop(0)
                                    df = pd.DataFrame(data, columns = cols)
                                    print(df.shape, cols)
                                    subset = [_conf["version"].get('field')]
                                    df.dropna(axis = 'rows', subset = subset, inplace = True)
                                    df[_conf["version"].get('field')] = np.vectorize(lambda field: str(field).upper())(df[_conf["version"].get('field')])
                                    version = df[_conf["version"].get('field')].max()
                                    new_map = _conf["map"].format(VERSION = str(version).upper())
                                map_version = new_map
                                if not new_map in mappin:
                                    try:                  
                                        sql = f"""SELECT * FROM "{map_version}" """
                                        mappin[map_version] = pd.read_sql(text(sql), engine.connect())
                                    except Exception as e:
                                        *_, exc_tb = sys.exc_info()
                                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                        return {
                                            'success': False,
                                            'msg': self.i18n('error-getin-map', _map = _input.get('spreadsheet_forms_map'))
                                        }
                            if map_has_versioning is True and map_version != '':
                                aux = self.spreadsheet_forms(file_pth, wb, mappin.get(map_version), file, self.params['file'], map_version, {})
                                if aux.get('success') is False:
                                    wb.close()
                                    return aux
                                elif aux.get('success') is True and aux.get('data'):
                                    proccessed_data.append(copy.deepcopy(aux['data']))
                            else:
                                for mp in mappin.items():
                                    aux = self.spreadsheet_forms(file_pth, wb, mappin[mp], file, self.params['file'], mp, {})
                                    if aux.get('success') is False:
                                        wb.close()
                                        return aux
                                    elif aux.get('success') is True and aux.get('data'):
                                        proccessed_data.append(copy.deepcopy(aux['data']))
                            wb.close()
                    except Exception as e:
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF: ', str(e), fname, exc_tb.tb_lineno)
                    zf.close()
                else:
                    pass
            elif _input.get('txt_fix_format_layout'):
                widths = _input.get('txt_fix_format_layout')
                is_ref = re.findall(r'^.+\[.+\]$', widths)
                if len(is_ref) > 0:
                    tbl = re.sub(r'\[.+\]$', '', widths)
                    fld = re.sub(r'\[|\]', '', widths)
                    print(widths, tbl, fld)
                    sql = text(f"""SELECT *\nFROM "{tbl}"\n""")
                    df_layout = pd.read_sql(sql , con = engine.connect())
                    widths = df_layout[fld].values
                else:
                    if len(re.findall(r'\;', widths)) > 0:
                        widths = widths.split(';')
                    elif len(re.findall(r'\,', widths)) > 0:
                        widths = widths.split(',')
                    elif len(re.findall(r'\|', widths)) > 0:
                        widths = widths.split(',')
                widths = [int(w) for w in widths]
                df = pd.read_fwf(f'{_path}/{fname}', widths = widths, names = headers, dtype = str, errors = 'coerce')
            elif ext.lower() in ['.csv'] or ext2.lower() in ['.csv']:         
                chunksize = 50 * 1000
                if _conf.get('chunksize'):
                    chunksize = _conf.get('chunksize')
                sep = ','
                if _conf.get('sep'):
                    sep = _conf.get('sep')
                else:
                    with open(f'{_path}/{fname}', mode = 'r', encoding = 'utf-8') as csv_file:
                        first_line = csv_file.readline()
                        if len(re.findall(r'\;', first_line)) > 0:
                            sep = ';'
                            if sep == ',' and len(re.findall(r'\|', first_line)) > 0:
                                sep = '|'
                df = pd.read_csv(f'{_path}/{fname}', chunksize = chunksize, usecols = columns_to_import, names = headers, skipinitialspace = True, sep = sep, low_memory=False)
            elif ext.lower() in ['.json']:
                df = pd.read_json(f'{_path}/{fname}')
            elif ext.lower() in ['.xlsx', '.xls', '.xlsm', '.xlsb']:
                if _conf.get('type') == 'excel-parts':
                    _excel_part_params = _conf.get('params')
                    df = pd.read_excel(
                        f'{_path}/{fname}',
                        sheet_name = _excel_part_params.get('sheet_name'),
                        skiprows = _excel_part_params.get('skiprows'), 
                        nrows = _excel_part_params.get('nrows'),
                        usecols = range(_excel_part_params.get('usecols')[0], _excel_part_params.get('usecols')[1])
                    )
                else:
                    engine = None
                    if ext.lower() in ['.xlsb']:
                        engine = 'pyxlsb'
                    multiple_sheets = _input.get('multiple_sheets', False)
                    specific_sheets = _input.get('specific_sheets', None)
                    if not specific_sheets:
                        pass
                    elif len(re.findall(r'\;', specific_sheets)) > 0:
                        specific_sheets = specific_sheets.split(';')
                    elif len(re.findall(r'\,', specific_sheets)) > 0:
                        specific_sheets = specific_sheets.split(',')
                    elif len(re.findall(r'\|', specific_sheets)) > 0:
                        specific_sheets = specific_sheets.split(',')
                    specific_range = _input.get('specific_range', None)
                    print(f'{_path}/{fname}')
                    if multiple_sheets is True:
                        xls = pd.ExcelFile(f'{_path}/{fname}')
                        df_sheets = []
                        for s in xls.sheet_names:
                            df_sheets.append(pd.read_excel(xls, s, usecols = columns_to_import, names = headers, engine = engine))
                        df = pd.concat(copy.deepcopy(df_sheets))
                    elif specific_sheets:
                        if isinstance(specific_sheets, str):
                            df = pd.read_excel(f'{_path}/{fname}', sheet_name = specific_sheets, usecols = columns_to_import, names = headers, engine = engine)
                            if df.get(specific_sheets):
                                df = df.get(specific_sheets)
                        elif len(specific_sheets) > 1:
                            df_aux = pd.read_excel(f'{_path}/{fname}', sheet_name = specific_sheets, usecols = columns_to_import, names = headers, engine = engine)
                            df = pd.concat([df_aux[d] for d in df_aux])
                            df_aux = None
                        else:
                            df = pd.read_excel(f'{_path}/{fname}', sheet_name = specific_sheets, usecols = columns_to_import, names = headers, engine = engine)
                            if df.get(specific_sheets):
                                df = df.get(specific_sheets)
                    else:
                        df = pd.read_excel(f'{_path}/{fname}', usecols = columns_to_import, names = headers, engine = engine)
            elif ext.lower() in ['.xlsb']:
                df = pd.read_excel(f'{_path}/{fname}', engine = 'pyxlsb', usecols = columns_to_import, names = headers)
            elif ext.lower() in ['.pq', '.parquet']:
                df = pd.read_parquet(f'{_path}/{fname}', engine = 'pyarrow')  
            elif ext.lower() in ['.db', '.sqlite', '.sqlite3']:
                url = URL.create(**{"drivername": "sqlite", "database": f'{_path}/{fname}'})
                engine = create_engine(url, echo = False)
                sql = f'SELECT * FROM "{_input.get("destination_table")}"'
                if _conf.get('query'):
                    sql = _conf.get('query')
                elif _conf.get('sql'):
                    sql = _conf.get('sql')             
                chunksize = 50 * 1000
                if _conf.get('chunksize'):
                    chunksize = _conf.get('chunksize')
                with engine.connect() as conn:
                    if engine.driver in ('pysqlite', 'sqlite'):
                        cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                        conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                        busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                        conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                    sql = self.set_query_date(sql, date_ref)
                    df = pd.read_sql(text(sql), con = conn, chunksize = chunksize)
            elif ext.lower() in ['.duckdb', '.ddb']:
                conn = duckdb.connect(f'{_path}/{fname}', read_only = False, config = {'memory_limit': '500mb'})
                sql = f'SELECT * FROM "{_input.get("destination_table")}"'
                if _conf.get('query'):
                    sql = _conf.get('query')
                elif _conf.get('sql'):
                    sql = _conf.get('sql')
                sql = _qd.set_date(sql, date_ref)
                df = conn.sql(sql).df()
                conn.close()   
            else:
                return {'success': False, 'msg': self.i18n('extention-not-suported', ext = ext)}   
            if isinstance(df, pd.core.frame.DataFrame):
                print(df.shape)
                return await self._run_import(df, _input, _etlrb, _conf, _conf_etlrb)
            elif isinstance(df, pd.io.parsers.readers.TextFileReader):
                _total = 0
                chunk = 0
                n_cols = 0
                for data in df:
                    _total += data.shape[0]
                    n_cols = data.shape[1]
                    chunk += 1
                    if chunk > 1:
                        _input['check_ref_date'] = False
                        _input['replace_existing_data'] = False
                    res = await self._run_import(data, _input, _etlrb, _conf, _conf_etlrb)
                    if res.get('success') is False:
                        return res
                return {'success': True, 'msg': self.i18n('success'), 'n_rows': _total, 'n_cols': n_cols}
            else:
                return {'success': False, 'msg': self.i18n('unexpected-error', err = 'ETL')}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # EXTRACT FROM A LIST OS XLSX SSFORMS USING DUCKDB
    async def _list_ssform_duckdb(self, _input, _etlrb, _conf, _conf_etlrb):
        '''Import a list of excel ssform files using duckdb
        {
            "type": "list-ssform-duckdb",
            "params": {
                "sql_fisrt": "SELECT * FROM \"SSFORM_LIST_TABLE\" ",
                "sql": """SELECT * FROM "SSFORM_LIST_TABLE" WHERE "FILE_FIELD" IS NOT NULL AND "AnexFicheiroContas" NOT IN (SELECT "filename" FROM "DEST_TABLE" WHERE "success" = 1)""",
                "field": "FILE_FIELD",
                "path": "static/uploads",
                "sql_extarct": """SELECT ANY_VALUE(TRIM("Field6")) FILTER(WHERE LOWER(TRIM("Field3")) LIKE LOWER('%name%of%the%field%in%ssform') AND "ROW" = 15) AS "name_of_the_field" FROM (SELECT ROW_NUMBER() OVER () AS "ROW", * FROM ST_READ('<filename>', layer = 'Bal')) AS "T""""
            }
        }'''
        try:
            _database = copy.deepcopy(self.conf.get('DATABASE'))
            _database['database'] = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
            if _conf.get('params'):
                _database = copy.deepcopy(_conf.get('params'))
            elif _conf_etlrb.get('params'):
                _database = copy.deepcopy(_conf_etlrb.get('params'))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
            else:
                engine = _db.get_engine(_database)
            if _conf['params'].get('sql'):
                try:
                    sql = _conf['params'].get('sql')
                    df = pd.read_sql(text(sql), con = engine.connect())
                except Exception as _err:# pylint: disable=broad-exception-caught
                    sql = _conf['params'].get('sql_fisrt')
                    df = pd.read_sql(text(sql), con = engine.connect())
                print('LIST OF XLS SSFROMS:', df.shape)
                sql_extarct = _conf['params'].get('sql_extarct')
                _field = _conf['params'].get('field')
                _path = _conf['params'].get('path')
                search_in_name = False
                _search_keys = []
                if _conf['params'].get('search_in_name') and isinstance(sql_extarct, dict):
                    search_in_name = True
                    _search_keys = list(sql_extarct.keys())
                err_if_not_sheet = _conf['params'].get('err_if_not_sheet', {})
                err_if_not_sheet_keys = err_if_not_sheet.keys()
                conn = duckdb.connect()
                conn.execute('INSTALL Spatial')
                conn.execute('LOAD Spatial')
                items = []
                for i,r in df.iterrows():
                    _row = dict(r)
                    if _row.get(_field):
                        fname = _row.get(_field)
                        _file_path = os.path.normpath(f'{os.getcwd()}/{_path}').encode("unicode_escape").decode("utf8")
                        if os.path.exists(f'{_file_path}/{fname}'):
                            #print(_file_path)
                            try:
                                _sql_to_run = None
                                if search_in_name and len(_search_keys) > 0:
                                    _selected_key = None
                                    for _key in _search_keys:
                                        _patt = re.compile(re.sub('_', '.+', _key), re.I)
                                        if len(re.findall(_patt, fname)) > 0:
                                            _selected_key = _key
                                            _sql_to_run = sql_extarct.get(_selected_key)
                                            break
                                    if not _selected_key:
                                        continue
                                        _row['extract_obs'] = f'No match for keys: {",".join(_search_keys)}'
                                        _row['success'] = False
                                    if not _sql_to_run:
                                        continue
                                        _row['extract_obs'] = f'No sql for key: {_selected_key}'
                                        _row['success'] = False                                        
                                else:
                                    _sql_to_run = sql_extarct
                                _sql = re.sub('<filename>', f'{_file_path}/{fname}', sql_extarct)
                                _sql = re.sub('<fname>', f'{_file_path}/{fname}', _sql)
                                try:
                                    _df = conn.sql(_sql).df()
                                except Exception as __err:
                                    if _selected_key and len(err_if_not_sheet_keys) > 0:
                                        for _sht in err_if_not_sheet_keys:
                                            if len(re.findall(_sht, str(__err)) > 0):
                                                _sql_to_run = err_if_not_sheet.get(_sht, {}).get(_selected_key)
                                                _sql = re.sub('<filename>', f'{_file_path}/{fname}', _sql_to_run)
                                                _sql = re.sub('<fname>', f'{_file_path}/{fname}', _sql)
                                                _df = conn.sql(_sql).df()
                                                break
                                #print({**dict(_df.loc[0])})
                                _row = {**_row, **dict(_df.loc[0])}
                                _row['extract_obs'] = ''
                                _row['success'] = True
                            except Exception as _err:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                print('DEBUG INF XLS SSFORMS: ', str(_err), self.params.get('ref'), fname, exc_tb.tb_lineno, f'{_file_path}/{fname}')
                                _row['extract_obs'] = str(_err)
                                _row['success'] = False
                                #continue
                        else: 
                            _row['extract_obs'] = 'File Not Found'
                            _row['success'] = False
                            #continue
                        items.append(copy.deepcopy(_row))
                conn.close()
            else:
                return {'success': False, 'msg': self.i18n('extract-failed')}
            df = pd.DataFrame(items)
            return await self._run_import(df, _input, _etlrb, _conf, _conf_etlrb)
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # EXTRACT FROM FILE DUCKD
    def set_str_env(self, _str: str) -> str:
        '''Set / replace environmental variable @ENV.NAME to real value in os.environ'''
        _patt = re.compile(r'@ENV\.\w+', re.I)
        match_env = re.findall(_patt, str(_str))
        if len(match_env) > 0:
            for _env in match_env:
                _environ = re.sub(r'@ENV\.', '', str(_env))
                try:
                    _str = re.sub(_env, os.environ.get(_environ), _str)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
        return _str
    async def _duckdb(self, _input, _etlrb, _conf, _conf_etlrb):
        '''Import data from file to duckdb
        ```json
            {
                "type": "file-duckdb",
                "duckdb": {
                    "database": "path 2 databse",
                    "register_filesystem": {
                        "fs": "ftp",
                        "args": {
                            "host": "@ENV.FTP_HOST",
                            "user": "@ENV.FTP_USR",
                            "password": "@ENV.FTP_PASS"
                        }
                    },
                    "pragmas_config_sql": [],
                    "extentions": ["spatial"],
                    "valid": [
                        {
                            "sql": "SELECT DISTINCT DATE_FIELD FROM '<filename>' WHERE DATE_FIELD = '{YYYYMMDD}' LIMIT 10",
                            "rule": "throw_if_not_empty",
                            "msg": "The DATE_FIELD in the '<filename>' is not YYYYMMDD!"
                        },
                        {
                            "sql": "SELECT DATE_FIELD FROM TBL WHERE DATE_FIELD = '{YYYYMMDD}' LIMIT 10",
                            "rule": "throw_if_empty",
                            "msg": "The date YYYYMMDD already exists, check the file or clean the data first!"
                        }
                    ],
                    "sql": "CREATE OR REPLACE TABLE TBL AS SELECT *, 'YYYY-MM-DD'::DATE AS "FILEREF", CURRENT_TIMESTAMP AS "IMPORTSTAMP" FROM '<filename>'",
                    "_sql": "INSERT INTO TBL AS SELECT * FROM db_scaner('<filename>', 'table_name')",
                    "__sql": "CREATE OR REPLACE TABLE TBL AS SELECT * FROM db_scan('<conn_string>', 'schema', 'table');"
                }
            }```'''
        try:
            fname = _input.get('file', '')
            is_tmp = _input.get('save_only_temp')
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}'
            if is_tmp is True:
                _path = tempfile.gettempdir()
            _path = os.path.normpath(_path).encode("unicode_escape").decode("utf8")
            basename, ext = os.path.splitext(fname)
            file_ref_patts = [
                {'patt': r'(\d{8})(?!.*\d+)', 'fmrt': '%Y%m%d'},
                {'patt': r'(\d{6})(?!.*\d+)', 'fmrt': '%Y%m'},
                {'patt': r'(\d{4})(?!.*\d+)', 'fmrt': '%y%m'}
            ]
            pats = {
                'file': re.compile(r'<filename>|<fname>|<file_name>|{filename}|{fname}|{file_name}', re.I),
                'table': re.compile(r'<table>|<table_name>|<tablename>|{table}|{table_name}|{tablename}', re.I)
            }
            if not _conf.get('duckdb'):
                _conf['duckdb'] = copy.deepcopy(_conf.get('params'))
            file_ref = None
            for patt in file_ref_patts:
                match = re.findall(patt.get('patt'), basename)
                if len(match) > 0:
                    try:
                        dt = datetime.datetime.strptime(str(match[0]), patt.get('fmrt'))                        
                        file_ref = datetime.date(dt.year, dt.month, calendar.monthrange(dt.year, dt.month)[1])
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        print(str(_err))
                    break
            if not file_ref or not date_ref:
                pass
            elif not _input.get('check_ref_date') or not _input.get('ref_date_field'):
                pass
            elif date_ref.strftime('%Y%m%d') != file_ref.strftime('%Y%m%d') and _input.get('ref_date_field') in ['file_ref', 'FILEREF', 'REF']:
                return {
                    'success': False, 
                    'msg': self.i18n('dt-file-no-match-dt-form', 
                    date_ref = date_ref.strftime('%Y-%m-%d'), 
                    file_ref = file_ref.strftime('%Y-%m-%d'))
                }
            database = _conf['duckdb'].get('database', _etlrb.get('database', _conf_etlrb.get('database')))
            #print({database})
            _db_path, _db_file = os.path.split(database)
            _db_basename, _db_ext = os.path.splitext(database)
            if not _db_path and _db_file:
                database = f'{os.getcwd()}/database/{database}'
            if not _db_ext and _db_basename:
                database = f'{database}.duckdb'
            # print(database)
            #url = URL.create(**{"drivername": "duckdb", "database": database})
            #engine = create_engine(url, echo = False)
            _qd = QueryDoc(query_parts = {})
            #conn = duckdb.connect(database, read_only = False, config = {'memory_limit': '500mb'})
            conn = duckdb.connect(database, read_only = False)
            if ext in ['.xlsx'] and not _conf['duckdb'].get('extentions'):
                _conf['duckdb']['extentions'] = ['spatial']
            # REGISTER FILESYSTEMS
            if not _conf['duckdb'].get('register_filesystem'):
                pass
            elif not isinstance(_conf['duckdb'].get('register_filesystem'), dict):
                pass
            elif not _conf['duckdb']['register_filesystem'].get('fs'):
                pass
            else:
                try:
                    #duckdb.register_filesystem(filesystem('gcs'))
                    fs = _conf['duckdb'].get('register_filesystem')
                    if fs.get('args'):
                        _args = fs.get('args', {})
                        for _key in _args:
                            _args[_key] = self.set_str_env(_args[_key])
                        conn.register_filesystem(filesystem(fs.get('fs'), **_args))
                    else:
                        conn.register_filesystem(filesystem(fs.get('fs')))
                except Exception as _err:
                    conn.close()
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF DUCKDB REGISTER FILESYSTEM: ', str(_err), fname, exc_tb.tb_lineno)
                    return {'success': False, 'msg': self.i18n('duckdb-register-filesystem-faild', fs = fs, err = str(_err))}
            # LOAD EXTENTIONS
            if not _conf['duckdb'].get('extentions'):
                pass
            elif isinstance(_conf['duckdb'].get('extentions'), list):
                for extention in _conf['duckdb'].get('extentions'):
                    try:
                        conn.install_extension(extention)
                        conn.load_extension(extention)
                    except Exception as _err:
                        conn.close()
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF DUCKDB EXTENTIONS: ', str(_err), fname, exc_tb.tb_lineno)
                        return {'success': False, 'msg': self.i18n('duckdb-extention-faild', extention = extention, err = str(_err))}
            # SET SOME PRAGMAS OR CONFIG PARAMETERS # ENV VARIABLES TO RUN BEFORE MAIN QUERY
            if not _conf['duckdb'].get('pragmas_config_sql_start'):
                pass
            elif isinstance(_conf['duckdb'].get('pragmas_config_sql_start'), list):
                for pragmas_config_sql_start in _conf['duckdb'].get('pragmas_config_sql_start'):
                    try:
                        pragmas_config_sql_start = self.set_str_env(pragmas_config_sql_start)
                        conn.sql(pragmas_config_sql_start)
                    except Exception as _err:
                        conn.close()
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF DUCKDB PRAGMA / CONFIG: ', str(_err), fname, exc_tb.tb_lineno)
                        return {'success': False, 'msg': self.i18n('duckdb-pragma-conf-faild', pconf = pragmas_config_sql_start, err = str(_err))}
            # ATACH THE DB IN CASE THE FILE IS ANTHER DUCKDB FILE
            if ext in ['.duckdb', '.ddb']:
                sql = f"ATTACH '<fname>' AS {basename}"
                sql = re.sub(pats['file'], f'{_path}/{fname}', sql)
                sql = _qd.set_date(sql, date_ref)
                #print('ATTACHING:', sql)
                try:
                    conn.sql(sql)
                except Exception as _err:
                    conn.close()
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF DUCKDB ATTACH: ', str(_err), fname, exc_tb.tb_lineno)
                    return {'success': False, 'msg': self.i18n('duckdb-atach-faild', sql = sql, err = str(_err))}
            # RUN VALIDATIONS QUERIES
            if not _conf['duckdb'].get('valid'):
                pass
            elif isinstance(_conf['duckdb'].get('valid'), list):
                for valid in _conf['duckdb'].get('valid'):
                    sql = valid.get('sql', valid.get('query'))
                    # print(f'{_path}/{fname}')
                    sql = re.sub(pats['file'], f'{_path}/{fname}', sql)
                    sql = re.sub(pats['table'], _input.get('destination_table'), sql)
                    sql = _qd.set_date(sql, date_ref)
                    sql = self.set_str_env(sql)
                    # print('DUCKDB VALIDATION: ', sql)
                    try:
                        df = conn.sql(sql).df()
                        rule = valid.get('rule')
                        msg = valid.get('msg', rule)
                        msg = _qd.set_date(msg, date_ref)
                        # print(msg, date_ref)
                        msg = re.sub(pats['file'], f'{fname}', msg)
                        msg = re.sub(pats['table'], _input.get('destination_table'), msg)
                        if rule == 'throw_if_not_empty' and df.shape[0] > 0:
                            return {'success': False, 'msg': msg}
                        elif rule == 'throw_if_empty' and df.shape[0] == 0:
                            return {'success': False, 'msg': msg}
                    except Exception as _err:
                        *_, exc_tb = sys.exc_info()
                        _fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF DUCKDB VALIDATION QUERIES: ', str(_err), _fname, exc_tb.tb_lineno)
                        _chk_first_time = re.findall(f"{_input.get('destination_table')}.+does not exist", str(_err))
                        if len(_chk_first_time) == 0:
                            conn.close()
                            return {'success': False, 'msg': self.i18n('duckdb-valid-faild', sql = sql, err = str(_err))}
            # EXECUTE THE MAIN QUERY
            if _conf['duckdb'].get('query'):
                sql = _conf['duckdb'].get('query')
                sql_bak = _conf['duckdb'].get('query')
            elif _conf['duckdb'].get('sql'):
                sql = _conf['duckdb'].get('sql')
                sql_bak = _conf['duckdb'].get('sql')
            sql = re.sub(pats['file'], f'{_path}/{fname}', sql)
            sql = re.sub(pats['table'], _input.get('destination_table'), sql)
            sql = _qd.set_date(sql, date_ref)
            sql = self.set_str_env(sql)
            # IF THE DATA IS GOIN TO BE APPENDED CHECK IF ALL THE COLUMNS EXISTS
            dtype = {
                'int8': 'INTEGER',
                'int32': 'BIGINT',
                'int64': 'BIGINT',
                'object': 'VARCHAR',
                'float64': 'DECIMAL',
                'float32': 'DECIMAL',
                'date': 'DATE',
                'datetime64[ns]': 'DATETIME',
                'datetime64': 'DATETIME',
                'bool': 'BOOLEAN'
            }
            patt = r'INSERT.+?INTO.+?[\"]?\w+[\"]?\.[\"]?.?\w+.?[\"]?|INSERT.+?INTO.+?.?\w+.?'
            patt = r'INSERT.+?INTO.+?.?\w+.?\.?.?.?\w+.?.?|INSERT.+?INTO.+?.?\w+.?'
            patt = r'INSERT.+?INTO.+[\"|\`|\[]]?\w+[\"|\`|\]]?\.[\"|\`|\[]]?.?\w+.?[\"|\`|\]]?|INSERT.+?INTO.+?.?\w+.?'
            patt = r'INSERT.+?INTO\s+?[\"|\`|\[]]?\w+[\"|\`|\]]?\.[\"|\`|\[]]?.?\w+.?[\"|\`|\]]?|INSERT.+?INTO\s+?[\"|\`|\[]]?.?\w+.?[\"|\`|\]]?'
            _match_insert_patt = re.findall(patt, re.sub(r'\n', ' ', sql))
            #print(r'INSERT.+?INTO.+?', _match_insert_patt)
            if len(_match_insert_patt) > 0:
                # print('THE QUERY IS TYPE INSERT / APPENDS DATA', _match_insert_patt[0])
                patt = r'INSERT.+?INTO.+?'
                table = re.sub(patt, '', _match_insert_patt[0])
                #  CHECK IF MATCHES "DB"."TABLE"
                patt = r'[\"|\`|\[]]?\w+[\"|\`|\]]?\.[\"|\`|\[]]?.?\w+.?[\"|\`|\]]?'
                _db = None
                if len(re.findall(patt, re.sub(r'\n', ' ', table))) > 0:
                    table = table.split('.')[1].strip()
                    _db = table.split('.')[0].strip()
                _table_striped = re.sub(r'"|\[|\]|\`|\s', '', table)
                sql_table_chk = f"""SELECT * 
                    FROM "information_schema"."tables" 
                    WHERE "table_name" = '{_table_striped}'"""
                if _db_basename:
                    # _db_basename              
                    sql_table_chk = f"""SELECT * 
                        FROM "information_schema"."tables" 
                        WHERE "table_name" = '{_table_striped}' 
                            AND LOWER("table_catalog") = LOWER('{_db_basename}')"""
                print(sql_table_chk)
                # CHECK IF THE TABLE EXISTS
                patt = r'INSERT.+?INTO.+?'
                _table_already_created = True
                try:
                    df = conn.sql(sql_table_chk).df()
                    if df.shape[0] == 0:
                        _table_already_created = False
                        #print(1, sql)
                        sql = re.sub(patt, 'CREATE TABLE ', sql)
                        if _conf['duckdb'].get('sql_create_if_not_exists'):
                            sql = _conf['duckdb'].get('sql_create_if_not_exists')
                        sql = re.sub(table, f'{table} AS ', sql, count = 1)
                        sql = re.sub(r'\bAS\s+AS\b', 'AS',sql)
                        print(2, sql)
                except Exception as _err:
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF DUCKDB CHECK IF TABLE EXISTS: ', table, str(_err), fname, exc_tb.tb_lineno)                     
                # FILE SCHEMA INFO 
                patt = r'SELECT.+'
                _sql_aux = re.findall(patt, re.sub(r'\n', ' ', sql))
                # CHECK IF IT MATCHES SELECT * FROM THE FILE / TABLE
                _match_select_star = []
                if len(_sql_aux) > 0:
                    patt = r'SELECT.+\*.+FROM\s+.?.?\w+|SELECT.+FROM\s+.?.?\w'
                    _match_select_star = re.findall(patt, _sql_aux[0])
                    #print('_match_select_star:', _match_select_star, '_table_already_created:', _table_already_created)
                if len(_match_select_star) == 1 and _table_already_created:
                    _file_columns_list = []
                    _file_columns_types = {}
                    if len(_sql_aux) > 0:
                        sql_column = f"SELECT * FROM ({_sql_aux[0]}) AS \"T\" LIMIT 10"
                        try:
                            df = conn.sql(sql_column).df()
                            _file_columns_list = df.columns
                            _file_columns_types = {column: str(df[column].dtype) for column in df.columns}
                        except Exception as _err:
                            *_, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            print('DEBUG INF DUCKDB GETTING FILE FIELDS: ', str(_err), fname, exc_tb.tb_lineno)
                    # DATABASE SCHEMA INFO 
                    patt = r'INSERT.+?INTO.+?'
                    table = re.sub(patt, '', _match_insert_patt[0])
                    #print(table)
                    _db_columns_list = []
                    #_db_columns_types = {}
                    sql_column = f"SELECT * FROM {table} LIMIT 10"
                    try:
                        df = conn.sql(sql_column).df()
                        _db_columns_list = df.columns
                        #print('_db_columns_list:', _db_columns_list)
                        #_db_columns_types = {column: str(df[column].dtype) for column in df.columns}
                    except Exception as _err:
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF DUCKDB GETTING FILE FIELDS: ', str(_err), fname, exc_tb.tb_lineno)                     
                    # CREATE COLUMN IN THE TABLE IF NOT EXISTS
                    if len(_db_columns_list) > 0 and len(_file_columns_list) > 0 and len(_db_columns_list) != len(_file_columns_list):
                        cols = '","'.join(_db_columns_list)
                        cols = f'"{cols}"'
                        # IN FILE BUT NOT IN THE DB
                        for column in _file_columns_list:
                            if column not in _db_columns_list:
                                column_type = _file_columns_types[column]
                                _sql = f"""ALTER TABLE {table} ADD "{column}" {dtype[column_type]} NULL"""
                                # print(column_type, _sql)
                                try:
                                    conn.sql(_sql)
                                    # _db_columns_list.append(column)
                                    cols = f'{cols},"{column}"'
                                except Exception as _err:
                                    *_, exc_tb = sys.exc_info()
                                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                    print('DEBUG INF DUCKDB ADD COLUMN FROM FILE: ', column, column_type, str(_err), fname, exc_tb.tb_lineno)
                        # IN THE DB BUT NOT IN FILE
                        for column in _db_columns_list:
                            if column not in _file_columns_list:
                                patt = re.compile(f'"{column}"')
                                cols = re.sub(patt, f'NULL AS "{column}"', f'{cols}')
                        # REPLACE SELECT * FOR SELECT **cols IN THE IMPORT STATMENT
                        #print(cols, sql_bak)
                        sql = re.sub(r'\*', cols, sql_bak)
                        sql = re.sub(pats['file'], f'{_path}/{fname}', sql)
                        sql = re.sub(pats['table'], _input.get('destination_table'), sql)
                        sql = _qd.set_date(sql, date_ref)
                        sql = self.set_str_env(sql)
            # RUN IMPORT QUERY
            print(sql)
            try:
                conn.sql(sql)
                n_rows = None
                # ROW COUNTS LOGIC
                patt = r'SELECT.+'
                _sql_aux = re.findall(patt, re.sub(r'\n', ' ', sql))
                if len(_sql_aux) > 0:
                    sql = f"SELECT COUNT(*) AS \"n_rows\" FROM ({_sql_aux[0]}) AS \"T\""
                    # print('TOTAL ROWS:', sql)
                    try:
                        df = conn.sql(sql).df()
                        n_rows = int(df['n_rows'][0])
                    except Exception as _err:
                        *_, exc_tb = sys.exc_info()
                        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                        print('DEBUG INF DUCKDB EXECUTE NUM ROWS: ', str(_err), fname, exc_tb.tb_lineno)
                # RUN SOME DETACHS IF EXISTS
                if not _conf['duckdb'].get('pragmas_config_sql_end'):
                    pass
                elif isinstance(_conf['duckdb'].get('pragmas_config_sql_end'), list):
                    for pragmas_config_sql_end in _conf['duckdb'].get('pragmas_config_sql_end'):
                        try:
                            pragmas_config_sql_end = self.set_str_env(pragmas_config_sql_end)
                            conn.sql(pragmas_config_sql_end)
                        except Exception as _err:
                            conn.close()
                            *_, exc_tb = sys.exc_info()
                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                            print('DEBUG INF DUCKDB PRAGMA / CONFIG: ', str(_err), fname, exc_tb.tb_lineno)
                            return {'success': False, 'msg': self.i18n('duckdb-pragma-conf-faild', pconf = pragmas_config_sql_end, err = str(_err))}
                conn.close()
                return {'success': True, 'msg': self.i18n('success'), 'n_rows': n_rows}
            except Exception as _err:
                conn.close()
                *_, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                print('DEBUG INF DUCKDB EXECUTE MAIN QUERY: ', str(_err), fname, exc_tb.tb_lineno)
                patt = re.compile(r'Invalid Input Error.+string.+to.+in column.+')
                if len(re.findall(patt, sql)) > 0:
                    _type_patt = re.compile(r'to (SMALLINT|INT1|INT|INTEGER|INT4|INT8|INT64|BIGINT|HUGEINT|LONG|DECIMAL|NUMERIC|DOUBLE|FLOAT8|REAL|FLOAT4|FLOAT|UBIGINT|UHUGEINT|UINTEGER|USMALLINT|UTINYINT|\s+) in')
                    _type = re.findall(_type_patt, sql)
                    print(1, _type)
                    if len(_type) > 0:
                        _type = _type[0]
                    _field_patt = re.compile(r'in.+column.+([\"|\`|\[]]?\w+[\"|\`|\]]?)\,')
                    _field = re.findall(_field_patt, sql)
                    print(2, _field)
                    if len(_field) > 0:
                        _field = _field[0]
                    _fix = None
                    if isinstance(_type, str) and isinstance(_field, str):
                        _fix = f"""REPLACE (TRY_CAST({_field} AS {_type}) AS {_field})"""
                    _sel_star_patt = re.compile(r'SELECT.+\*.+FROM')                    
                    if len(re.findall(_sel_star_patt, sql)) > 0:
                        patt = re.compile(r'SELECT.+\*.+(REPLACE.?\(.+\)).+FROM')
                        has_replace = re.findall(patt, sql)
                        if len(has_replace) > 0:
                            has_replace = has_replace[0]
                            patt = re.compile(r'\)$')
                            has_replace = re.sub(patt, f', {_fix})', has_replace)
                            sql = re.sub(_sel_star_patt, f'SELECT * {has_replace} FROM', sql)
                        else:
                            sql = re.sub(_sel_star_patt, f'SELECT * REPLACE({_fix}) FROM', sql)
                        print(3,sql)
                        if _conf['duckdb'].get('query'):
                            _conf['duckdb']['query'] = sql
                        elif _conf['duckdb'].get('sql'):
                            _conf['duckdb']['sql'] = sql
                        self._duckdb(_input, _etlrb, _conf, _conf_etlrb)
                return {'success': False, 'msg': self.i18n('duckdb-import-faild', sql = sql, err = str(_err))}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF DUCKDB: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # EXTRACT FROM DB
    async def _get_new_db(self, _database):
        '''get db connection'''
        if isinstance(_database, dict):
            _patt = re.compile(r'@ENV\..+')
            for _key in _database:
                match_env = re.findall(_patt, str(_database[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _database[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            patt = re.compile(r"\{.*?\}", re.IGNORECASE)
            matchs_odbc = re.findall(patt, _database.get('drivername'))
            if len(matchs_odbc) > 0:
                db_api = _database.get('odbc_dbapi')
                _prms = _database.get('odbc_url').format(**_database)
                #print(_prms)
                prms = urllib.parse.quote_plus(_prms)
                return create_engine(f"{db_api}+pyodbc:///?odbc_connect={prms}")
            else:
                # self.params['database'] = _database
                return DB(self.conf, self.params)
        return DB(self.conf, self.params)
    async def _from_db(self, _input, _etlrb, _conf, _conf_etlrb):
        '''extract from db'''
        try:
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            _database = copy.deepcopy(_conf.get('params'))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
                metadata = sa.MetaData()
            else:
                engine = _db.get_engine(_database)
                metadata = self.db.get_metadata(engine)
            # print('ORIGN:', engine.url)
            incremental_extract = _input.get('incremental_extract', False)
            last_update_date_field = _input.get('last_update_date_field', None)
            ref_id_keys = _input.get('ref_id_keys', None)
            if incremental_extract is True and last_update_date_field and ref_id_keys:
                _database_dest = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
                _db_dest = await self._get_new_db(_database_dest)
                if isinstance(_db_dest, sa.engine.base.Engine):
                    engine_dest = _db_dest
                    metadata_dest = sa.MetaData()
                else:
                    engine_dest = _db_dest.get_engine(_database_dest)
                    metadata_dest = self.db.get_metadata(engine_dest)
                destination_table = _input.get('destination_table')
                sa_table_dest = None
                try:
                    sa_table_dest = Table(destination_table, metadata_dest, autoload_with = engine_dest)
                except Exception as _err:# pylint: disable=broad-exception-caught
                    pass
                last_dt_dest = None
                if sa_table_dest is not None:
                    # pylint: disable=not-callable
                    sql = select(func.max(sa_table_dest.c[last_update_date_field]))\
                            .select_from(sa_table_dest)
                    conn_dest = engine_dest.connect()
                    if engine.driver in ('pysqlite', 'sqlite'):
                        cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                        conn_dest.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                        busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                        conn_dest.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                    res = conn_dest.execute(sql).fetchone()
                    last_dt_dest = res['max_1']
                    #print(last_dt_dest, type(last_dt_dest), sql)
                    if last_dt_dest:
                        _org_table = _conf.get('org_table') if _conf.get('org_table') else destination_table
                        sa_table_org = Table(_org_table, metadata, autoload_with = engine)
                        _dt_aux = last_dt_dest
                        if isinstance(last_dt_dest, datetime.datetime):
                            _dt_aux = last_dt_dest.strftime('%Y-%m-%d %H:%M:%S')
                        sql = select(sa_table_org)\
                                .select_from(sa_table_org)\
                                .where(sa_table_org.c[last_update_date_field] > _dt_aux)
                        conn = engine.connect()
                        if engine.driver in ('pysqlite', 'sqlite'):
                            cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                            conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                            busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                            conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                        result = conn.execute(sql)
                        res = result.mappings().all()
                        result.close()
                        conn.close()
                        _data = []
                        for row in res:
                            d = {}
                            for column in row:
                                d[column] = row[column]
                            _data.append(d)
                        _ids = list(map(lambda d: d.get(ref_id_keys), _data))
                        sql = select(sa_table_dest.c[ref_id_keys])\
                                .select_from(sa_table_dest)\
                                .where(sa_table_dest.c[ref_id_keys].in_(_ids))
                        res = conn_dest.execute(sql)
                        _ids_dest = list(map(lambda d: d.get(ref_id_keys), res.fetchall()))
                        # print(_ids, _ids_dest)
                        res.close()
                        for d in _data:
                            if d.get(ref_id_keys) in _ids_dest:
                                sql = sa.update(sa_table_dest)\
                                        .where(sa_table_dest.c[ref_id_keys] == d.get(ref_id_keys))\
                                        .values(d)
                            else:
                                sql = sa.insert(sa_table_dest).values(d)
                            conn_dest.execute(sql)
                        conn_dest.close()
                        return {'success': True, 'msg': self.i18n('success'), 'n_rows': len(_ids)}           
            sql = f"""SELECT * FROM "{_input.get('destination_table')}";"""
            if _conf.get('query'):
                sql = _conf.get('query')
            elif _conf.get('sql'):
                sql = _conf.get('sql')
            sql = self.set_query_date(sql, date_ref)
            # print(sql, engine.url, _database)
            #print(sql)
            chunksize = 50 * 1000
            if _conf.get('chunksize'):
                chunksize = _conf.get('chunksize')
            reader = pd.read_sql(text(sql), con = engine.connect(), chunksize = chunksize, coerce_float = True)
            _total = 0
            chunk = 0
            n_cols = 0
            for data in reader:
                _total += data.shape[0]
                n_cols = data.shape[1]
                chunk += 1
                if chunk > 1:
                    _input['check_ref_date'] = False
                    _input['replace_existing_data'] = False
                res = await self._run_import(data, _input, _etlrb, _conf, _conf_etlrb)
                if res.get('success') is False:
                    return res
            return {'success': True, 'msg': self.i18n('success'), 'n_rows': _total, 'n_cols': n_cols}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # _from_odbc_csv_duckdb
    async def _from_odbc_csv_duckdb(self, _input, _etlrb, _conf, _conf_etlrb):
        '''extract from odbc to csv then csv file to duckdb via file-duckdb'''
        try:
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            sql = f"""SELECT * FROM "{_input.get('destination_table')}";"""
            if _conf.get('query'):
                sql = _conf.get('query')
            elif _conf.get('sql'):
                sql = _conf.get('sql')
            sql = self.set_query_date(sql, date_ref)
            conn_str = self.set_str_env(_conf.get('params', {}).get('odbc_conn'))
            #print(conn_str, sql)
            #bin_path = r'rust-odbc-csv'
            #res = subprocess.check_output([bin_path, conn_str, sql.strip(), '50000', _input.get('destination_table')])
            res = odbc_csv([conn_str, sql])
            res_json = json.loads(res)
            #print(res_json)
            if res_json.get('success') is True and res_json.get('fname'):
                #print(res_json)
                _input['file'] = os.path.split(res_json.get('fname'))[1]
                _input['save_only_temp'] = True
                return await self._duckdb(_input, _etlrb, _conf, _conf_etlrb)
            return res_json 
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF ODBC-CSV-DUCDB: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # FTP
    async def _from_ftp(self, _input, _etlrb, _conf, _conf_etlrb):
        '''extract ftp'''
        try:
            _patt = re.compile(r'@ENV\..+')
            _params = _conf['params']
            for _key in _params:
                match_env = re.findall(_patt, str(_params[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _params[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = parser.parse(_input.get('date_ref'))
            fil = self.set_query_date(_params.get('file'), date_ref)
            ftp = FTP(_params.get('host', _params.get('hostname')))
            ftp.login(
                user = _params.get('username', _params.get('user')), 
                passwd = _params.get('password', _params.get('passwd'))
            )
            ftp.cwd(_params.get('cwd'))
            basename, ext = os.path.splitext(fil)
            #print(basename, ext)
            fil2 = re.sub(r'00000\d$', '', fil)
            if not ext:
                fil2 = f'{fil2}.txt'
            if _params.get('ext_overwrite'):
                fil2 = f"{fil2}.{_params.get('ext_overwrite')}"   
            path = f'{tempfile.gettempdir()}/{fil2}'
            # print(path, fil)
            file_writer = open(path, 'wb')
            ftp.retrbinary(f'RETR {fil}', file_writer.write)
            file_writer.close()
            ftp.quit()
            _input['file'] = fil2
            _input['save_only_temp'] = True
            return await self._from_file(_input, _etlrb, _conf, _conf_etlrb)
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # HTTP
    async def _from_webscrap(self, _input, _etlrb, _conf, _conf_etlrb):
        '''# extract webscrap
        input config:
        ```json
            {
                "type": "webscrap",
                "params": {
                    "url": str,
                    "table_id": str,
                    "timeout": int,
                    "columns": list[str]
                }
            }
        ```
        '''
        try:
            _patt = re.compile(r'@ENV\..+')
            _params = _conf['params']
            for _key in _params:
                match_env = re.findall(_patt, str(_params[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _params[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            url = _params.get('url', _params.get('link'))
            table_id = _params.get('table_id')
            headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"}
            timeout = _params.get('timeout', 60)
            response = requests.get(url, headers = _params.get('headers', headers), timeout = timeout)
            soup = BeautifulSoup(response.content, "html.parser")
            table = soup.find("table", id = table_id)
            _df = pd.read_html(str(table), flavor="bs4")[0]
            if _params.get('columns', _params.get('names')):
                _df.columns = _params.get('columns', _params.get('names'))        
            return await self._run_import(_df, _input, _etlrb, _conf, _conf_etlrb)
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # WEB MAIL
    async def _from_outlook(self, _input, _etlrb, _conf, _conf_etlrb):
        '''# extract outlook
        input config:
        ```json
            {
                "type": "outlook_mail",
                "params": {
                    "return_data": false,
                    "subject_filter": "TASK_NAME",
                    "body_filter": null,
                    "sender_filter": null,
                    "sender_email_filter": null,
                    "recipients_filter": null,
                    "recipients_email_filter": null,
                    "date_filter": true,
                    "save_attachemt": true,
                    "main_folder": ["user@domai.com"],
                    "folders": ["A Receber", "Inbox"],
                    "patt_title_match": "TASK_NAME.\\d{4}.?\\d{2}.?\\d{2}",
                    "patt_title_exclude": "Não.+entregue.+|Nao.+entregue.+|Not.+delivered",
                    "fields": [
                        {
                            "name": "date_ref",
                            "type": "date",
                            "patt": "\\d{4}.?\\d{2}.?\\d{2}",
                            "where": "subject",
                            "format": "%Y%m%d"
                        }
                    ]
                }
            }
        ```
        '''
        try:
            _patt = re.compile(r'@ENV\..+')
            _params = _conf['params']
            for _key in _params:
                match_env = re.findall(_patt, str(_params[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _params[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            date_ref = parser.parse(_input.get('date_ref'))
            #print(_params)
            #PROCESS EMAILS
            patt_title_match = re.compile(_params.get('patt_title_match'), re.IGNORECASE)
            #print(patt_title_match)
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}/tmp'
            items = []
            try:
                #outlook  = win32com.client.Dispatch("Outlook.Application", pythoncom.CoInitialize()).GetNamespace("MAPI")
                outlook  = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                for folder in outlook.Folders:
                    # MAIN FOLDER
                    #print(1, folder.Name)
                    if not _params.get('main_folder'):
                        continue
                    elif folder.Name not in _params.get('main_folder') and str(folder.Name).lower() not in _params.get('main_folder'):
                        continue
                    recip = outlook.CreateRecipient(str(folder))
                    # SUBFORDERS
                    for sub_folder in  folder.Folders:
                        #print(2, sub_folder.Name)
                        if not _params.get('folders'):
                            continue
                        elif sub_folder.Name not in _params.get('folders') and sub_folder.Name.lower() not in _params.get('folders'):
                            continue
                        # LOOP MSG's
                        messages = sub_folder.Items
                        if _params.get('subject_filter'):
                            _filter = _params.get('subject_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:subject\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('body_filter'):
                            _filter = _params.get('body_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:textdescription\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('sender_email_filter'):
                            _filter = _params.get('sender_email_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:fromemail\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('sender_filter'):
                            _filter = _params.get('sender_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:fromname\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('recipients_email_filter'):
                            _filter = _params.get('recipients_email_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:displayto\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('recipients_filter'):
                            _filter = _params.get('recipients_filter')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:displayto\" ci_phrasematch '{_filter}'"
                            messages = messages.Restrict(filter_str)
                            #print(filter_str)
                        if _params.get('date_filter'):
                            _date = date_ref.strftime('%Y-%m-%d')
                            filter_str = f"@SQL=\"urn:schemas:httpmail:datereceived\" >= '{_date}'"
                            #print(filter_str)
                            messages = messages.Restrict(filter_str)
                        message  = messages.GetFirst()
                        while message:
                            try:
                                d = {}
                                d['subject'] = str(getattr(message, 'Subject', '<UNKNOWN>'))
                                #print(d['subject'])
                                try:
                                    d['sent_on']  = getattr(message, 'SentOn', '<UNKNOWN>').strftime('%Y-%m-%d %H:%M:%S')
                                except Exception as _err: # pylint: disable=broad-exception-caught
                                    d['sent_on']  = str(getattr(message, 'SentOn', '<UNKNOWN>'))
                                    #continue
                                # d['EntryID'] = getattr(message, 'EntryID', '<UNKNOWN>')
                                d['sender']  = str(getattr(message, 'Sender', '<UNKNOWN>'))
                                d['size']    = str(getattr(message, 'Size', '<UNKNOWN>'))
                                d['body']    = str(getattr(message, 'Body', '<UNKNOWN>'))
                                d['body']    = re.sub(r'[\n\r]{2,}', '\n', d['body'])
                                d['body']    = re.sub(r'[\n\r]{2,}', '\n', d['body'])
                                #print(d['Body'])
                                #break
                                match = re.findall(patt_title_match, d['subject'])
                                match_exclude = []
                                if _params.get('patt_title_exclude'):
                                    patt_title_exclude = re.compile(_params.get('patt_title_exclude'), re.IGNORECASE)
                                    match_exclude = re.findall(patt_title_exclude, d['subject'])
                                if len(match) > 0 and len(match_exclude) == 0:
                                    recipients = message.Recipients
                                    _recipients = []
                                    for recipient in recipients:
                                        _recipients.append(recipient.Name)
                                    d['recipients'] = ';'.join(_recipients)
                                    if _params.get('save_attachemt'):
                                        try:
                                            _fname = f'{_path}/{secure_filename(d["subject"])}.msg'
                                            if not os.path.exists(_fname):
                                                message.SaveAs(_fname, 3)
                                            d['fname'] = f'tmp/{secure_filename(d["subject"])}.msg'
                                        except Exception as _err:
                                            print(str(_err))
                                    # print(folder.Name, sub_folder.Name, d['subject'])
                                    for field in _params.get('fields'):
                                        #print(field)
                                        d[field['name']] = ''
                                        patt = re.compile(field.get('patt'), re.IGNORECASE)
                                        match = re.findall(patt, d[field.get('where')])
                                        if len(match) > 0:
                                            d[field['name']] = match[0]
                                            if field.get('replace'):
                                                d[field['name']] = re.sub(re.compile(field['replace'][0]),  field['replace'][1],str(d[field['name']]))
                                            d[field['name']] = re.sub(r'\r', ' ', d[field['name']])
                                            d[field['name']] = re.sub(r'\t', ' ', d[field['name']])
                                            d[field['name']] = re.sub(r'\n', ' ', d[field['name']])
                                            d[field['name']] = re.sub(r';', '', d[field['name']])
                                            if field.get('patt2') and d.get(field['name']):
                                                patt2 = re.compile(field.get('patt2'), re.IGNORECASE)
                                                match2 = re.findall(patt2, d.get(field['name']))
                                                # print(patt2, match2)
                                                if len(match2) > 0:
                                                    d[field['name']] = match2[0]
                                            if field.get('format') and field.get('type') == 'date' and d.get(field['name']):
                                                try:
                                                    d[field['name']] = re.sub(r'\s', '', d[field['name']])
                                                    d[field['name']] = datetime.datetime.strptime(d[field['name']], field.get('format')).strftime('%Y-%m-%d')
                                                except Exception as e:
                                                    print(str(e), field['name'], d[field['name']])
                                    d['main_folder']  = folder.Name
                                    d['sub_folder']  = sub_folder.Name
                                    #print(d)
                                    items.append(dict(d))
                            except Exception as inst:
                                exc_type, exc_obj, exc_tb = sys.exc_info()
                                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                print('DEBUG INF: ', str(inst), fname, exc_tb.tb_lineno)
                            message = messages.GetNext()
                #outlook.close()
            except Exception as e:
                print(str(e))
            if _params.get('return_data') is True:
                return {'success': True, 'msg': self.i18n('success'), 'data': items}
            _df = pd.DataFrame(items)       
            return await self._run_import(_df, _input, _etlrb, _conf, _conf_etlrb)
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # spreadsheet forms
    def spreadsheet_forms(self, path, wb, mappin, file, pakIfZip, map_name, other_props):
        '''spreadsheet forms'''
        try:
            if not wb:
                wb = xl.load_workbook(path, read_only = True, keep_links = False, data_only = True)
            sheets = wb.sheetnames
            for sh in sheets:
                wb[sh].title = wb[sh].title.strip()
            sheets = wb.sheetnames
            sheetsMapped = mappin['SHEET'].unique()
            for sh in sheetsMapped:
                if sh not in sheets:
                    break
            has_at_least_one_seet = False
            for sh in sheetsMapped:
                if sh in sheets:
                    has_at_least_one_seet = True
            if has_at_least_one_seet is False:
                _msg = '''Error processing the file, in "{mapp}" is expected these sheets "{mappinSheets}", but got "{fileSheets}"!'''
                return {
                    'success': False,
                    'msg': _msg.format(
                        mapp = map_name,
                        mappinSheets = '", "'.join(sheetsMapped),
                        fileSheets = '", "'.join(sheets)
                    )
                }
            aux = {}
            selected_sheet = ''
            for i,r in mappin.iterrows():
                try:
                    if selected_sheet != r['SHEET']:
                        selected_sheet = r['SHEET']
                        ws = wb[selected_sheet]
                    if ws.sheet_state != 'visible':
                        if f'{selected_sheet} ' in wb.sheetnames:
                            ws = wb[f'{selected_sheet} ']
                            if ws.sheet_state != 'visible':
                                print(selected_sheet, ws.sheet_state, file)
                                continue
                        else:
                            continue
                    aux[r['COLUMN'].strip()] = ws[r['REF']].value
                    try:
                        if pd.isna(float(aux[r['COLUMN'].strip()])):
                            aux[r['COLUMN'].strip()] = None
                    except Exception as _err:
                        pass
                except Exception as e:
                    *_, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF: ', file, r['SHEET'], r['COLUMN'].strip(), r['REF'], str(e), fname, exc_tb.tb_lineno)
                    if str(e).find('date value out of range') != -1:
                        aux[r['COLUMN'].strip()] = ''
                    else:
                        aux[r['COLUMN'].strip()] = ''
            if len(aux) > 0:
                prop = wb.properties
                aux['mapp'] = map_name
                # aux['DATA_IMP'] = dt_aux.datetime.now()
                aux['file'] = file
                nums = re.findall(re.compile(r'\d{3,}'), file)
                posbl_ncli = []
                if len(nums) > 0:
                    for n in nums:
                        try:
                            if len(str(int(n))) > 2:
                                posbl_ncli.append(n)
                        except Exception as e:
                            pass
                    if len(posbl_ncli) > 0:         
                        aux['NUM'] = posbl_ncli[0]
                aux['zip_file'] = pakIfZip
                aux['creator'] = prop.creator
                aux['created'] = prop.created
                aux['modified'] = prop.modified
                if not other_props:
                    pass
                elif other_props.get('Last save time'):
                    try:
                        aux['modified'] = other_props.get('Last save time').strftime('%Y-%m-%d %H:%M:%S')
                    except Exception as e:
                        print('Err Last save time', str(e))
                aux['last_modified_by'] = prop.last_modified_by
                if not other_props:
                    pass
                elif other_props.get('Last author'):
                    aux['last_modified_by'] = other_props.get('Last author')
                return {
                    'success': True,
                    'data': aux
                }
            else:
                wb.close()
                print(aux, path, file, map_name)
                return {
                    'success': False,
                    'msg': self.i18n('err-file-proc-no results'),
                    '_msg': 'Erro ao processar o ficheiro, o resultado retornou vazio, certifique o ficheiro selecionado!'
                }
        except Exception as _err:
            wb.close()
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # SET DATA TO QUIERY STRING
    def _get_dt_fmrt(self, _format):
        '''return python date format'''
        _py_format = _format
        _frmts = [
            {'frmt': r'YYYY|AAAA', 'py_fmrt': '%Y'},
            {'frmt': r'YY|AA', 'py_fmrt': '%y'},
            {'frmt': r'MM', 'py_fmrt': '%m'},
            {'frmt': r'DD', 'py_fmrt': '%d'}
        ]
        for _fmrt in _frmts:
            _py_format = re.sub(re.compile(_fmrt['frmt'], re.I), _fmrt['py_fmrt'], _py_format)
        return _py_format
    def set_query_date(self, query, date_ref):
        '''set query date'''
        patt = re.compile(r"([\"]?\w+[\"]?\.[\"]?\w+[\"]?\s{0,}=\s{0,}'\{.*?\}'|[\"]?\w+[\"]?\s{0,}=\s{0,}'\{.*?\}')", re.IGNORECASE)
        matchs = re.findall(patt, query)
        if not matchs:
            patt = re.compile(r"[\"]?\w+[\"]?\s{0,}=\s{0,}'\{.*?\}'", re.IGNORECASE)
            matchs = re.findall(patt, query)
        elif len(matchs) == 0:
            patt = re.compile(r"[\"]?\w+[\"]?\s{0,}=\s{0,}'\{.*?\}'", re.IGNORECASE)
            matchs = re.findall(patt, query)
        if len(matchs) > 0:
            patt2 = re.compile(r"'\{.*?\}'", re.IGNORECASE)
            for m in matchs:
                frmt = re.findall(patt2, m)
                if len(frmt) > 0:
                    frmt_final = self._get_dt_fmrt(frmt[0])
                    frmt_final = frmt_final.replace('{','').replace('}','')
                    if isinstance(date_ref, list):
                        dts = ','.join([f'{dt.strftime(frmt_final)}' for dt in copy.deepcopy(date_ref)])
                        procc = re.sub(patt2, f'({dts})', m)
                        patt3 = re.compile(r"\s?=\s?", re.IGNORECASE)
                        procc = re.sub(patt3, ' IN ', procc)
                    else:
                        procc = re.sub(patt2, date_ref.strftime(frmt_final), m)
                    patt =  re.compile(r'(' + m + ')', re.IGNORECASE)
                    query = re.sub(patt, procc, query)
        patt = re.compile(r"'?\{.*?\}'?", re.IGNORECASE)
        matchs = re.findall(patt, query)
        if len(matchs) > 0:
            for m in matchs:
                frmt = m
                frmt_final = self._get_dt_fmrt(frmt)
                frmt_final = frmt_final.replace('{','').replace('}','')
                if isinstance(date_ref, list):
                    procc = re.sub(patt, date_ref[0].strftime(frmt_final), m)
                else:
                    procc = re.sub(patt, date_ref.strftime(frmt_final), m)
                patt =  re.compile(r'(' + m + ')', re.IGNORECASE)
                query = re.sub(patt, procc, query)
        # IN CASE WE DO HAVE TEMP TABLES WITH DATE EXTENTIONS 
        patt = re.compile(
            r'YYYY.?MM.?DD|AAAA.?MM.?DD|YY.?MM.?DD|AA.?MM.?DD|YYYY.?MM|AAAA.?MM|YY.?MM|AA.?MM|MM.?DD|DD.?MM.?YYYY|DD.?MM.?AAAA|DD.?MM.?YY|DD.?MM.?AA'
            , re.IGNORECASE
        )
        matchs = re.findall(patt, query)
        if len(matchs) > 0:
            for m in matchs:
                frmt = m
                frmt_final = self._get_dt_fmrt(frmt)
                # frmt_final = frmt_final.replace('{','').replace('}','')
                if isinstance(date_ref, list):
                    procc = re.sub(patt, date_ref[0].strftime(frmt_final), m)
                else:
                    procc = re.sub(patt, date_ref.strftime(frmt_final), m)
                patt =  re.compile(r'(' + m + ')', re.IGNORECASE)
                query = re.sub(patt, procc, query)      
        return query
    # DELETE
    async def delete(self):
        '''DELETE'''
        try:
            _data = self.params['data']
            _input = _data['data']
            _etlrb = _data.get('selected_etlrb')
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = [parser.parse(_input.get('date_ref'))]
            _database = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
                metadata = sa.MetaData()
            else:
                engine = _db.get_engine(_database)
                metadata = self.db.get_metadata(engine)
            destination_table = _input.get('destination_table')
            ref_date_field = _input.get('ref_date_field', None)
            sa_table = Table(destination_table, metadata, autoload_with = engine)
            with engine.connect() as conn:
                if engine.driver in ('pysqlite', 'sqlite'):
                    cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                    conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                    busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                    conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                if ref_date_field:
                    _dt_frmt = '%Y-%m-%d'
                    if _input.get('date_format_org'):
                        _dt_frmt = self._get_dt_fmrt(_input.get('date_format_org'))
                        #print(_input.get('ref_date_field'), _input.get('date_format_org'), _dt_frmt)
                    _dts = [d if isinstance(d, str) else d.strftime(_dt_frmt) for d in date_ref]
                    sql = sa.delete(sa_table)\
                            .where(sa_table.c[ref_date_field].in_(_dts))
                else:
                    sql = sa.delete(sa_table)
                print((sql))
                res = conn.execute(sql)
                conn.commit()
                n_rows = res.rowcount
                res.close()
                conn.close()
                return {'success': True, 'msg': self.i18n('success'), 'n_rows': n_rows}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}    
    # N_ROWS
    async def n_rows(self):
        ''' returns the number of records '''
        try:
            _data = self.params['data']
            _input = _data['data']
            _etlrb = _data.get('selected_etlrb')
            date_ref = _input.get('date_ref')
            if isinstance(_input.get('date_ref'), list):
                date_ref = [parser.parse(d) if isinstance(d, str) else d for d in _input.get('date_ref')]
            elif isinstance(_input.get('date_ref'), str):
                date_ref = [parser.parse(_input.get('date_ref'))]
            _database = copy.deepcopy(self.params.get('database', self.params.get('db', _etlrb.get('database', None))))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
                metadata = sa.MetaData()
            else:
                engine = _db.get_engine(_database)
                #print(2, engine.url)
                metadata = _db.get_metadata(engine)
            destination_table = _input.get('destination_table')
            ref_date_field = _input.get('ref_date_field', None)
            #print(destination_table, type(metadata), engine.url)
            sa_table = Table(destination_table, metadata, autoload_with = engine)
            with engine.connect() as conn:
                if engine.driver in ('pysqlite', 'sqlite'):
                    cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                    conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                    busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                    conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                if ref_date_field:
                    _dt_frmt = '%Y-%m-%d'
                    if _input.get('date_format_org'):
                        _dt_frmt = self._get_dt_fmrt(_input.get('date_format_org'))
                        #print(_input.get('ref_date_field'), _input.get('date_format_org'), _dt_frmt)
                    _dts = [d if isinstance(d, str) else d.strftime(_dt_frmt) for d in date_ref]
                    # pylint: disable=not-callable
                    sql = select(func.count(sa_table.c[ref_date_field]))\
                        .select_from(sa_table)\
                        .where(sa_table.c[ref_date_field].in_(_dts))
                else:
                    # pylint: disable=not-callable
                    sql = select(func.count(text('*'))).select_from(sa_table)# pylint: disable=not-callable
                res = conn.execute(sql)
                data = res.fetchone()
                n_rows = data[0]
                # print(sql, [d if isinstance(d, str) else d.strftime('%Y-%m-%d') for d in date_ref], data)
                res.close()
                conn.close()
                return {'success': True, 'msg': self.i18n('success'), 'n_rows': n_rows}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
