'''EXPOTS'''
# pylint: disable=unused-variable
# pylint: disable=line-too-long
# pylint: disable=wrong-import-order
# pylint: disable=broad-exception-caught
# pylint: disable=unused-import
# pylint: disable=invalid-name
# pylint: disable=trailing-whitespace
import os
import sys
import tempfile
import datetime
import copy
import re
import json
from dateutil import parser
import pandas as pd
import numpy as np
import pdfkit
import sqlalchemy as sa
from sqlalchemy import create_engine, text
import urllib
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from jinja2 import Environment, BaseLoader#, FileSystemLoader, select_autoescape # , PackageLoader
from core.crud import Crud
from core.db import DB
from core.export_duckdb import ExportDucdb
from query_doc import QueryDoc
#env = Environment(loader = FileSystemLoader(searchpath="./", encoding='utf-8'), autoescape = select_autoescape(['html', 'xml', 'tpl']))


class Export:
    '''EXPOTS'''
    def __init__(self, conf, params, db, i18n):
        self.conf = conf
        self.params = params
        self.db = db
        self.i18n = i18n
    # EXPORT CONTENT FROM THE READ API
    async def read(self):
        '''dump from CRUD.read method'''
        try:
            _conf = self.params['data'].get('_conf')
            # IF USER CHECKED ALL RECORDAS
            if _conf.get('records') == 'all_records':
                self.params['data']['limit'] = -1
            # print(_conf, _conf.get('_fields'))
            _crud = Crud(self.conf, self.params, self.db, self.i18n)
            _read = await _crud.read()
            if not _read.get('success'): # IN CASE OF ERR
                return _read
            else:
                df = pd.DataFrame(_read.get('data'))
                fields = []
                # CHOOSES THE FIELDS TO DUMP
                if _conf.get('display_fields') == 'interface_fields'\
                    and self.params['data'].get('_fields'):
                    _fields = filter(
                        lambda field: field.get('display') is True,
                        self.params['data'].get('_fields')
                    )
                    fields = list(map(
                        lambda field: field.get('name'),
                        list(_fields)
                    ))
                    df = df[fields]
                    #print(df.shape, df.columns)
                # COLUMNS NAMES
                if _conf.get('column_names') == 'comments'\
                    and self.params['data'].get('_fields'):
                    new_names = []
                    aux_names_dict = {}
                    for field in self.params['data'].get('_fields'):
                        aux_names_dict[field['name']] = field
                    for c in df.columns:
                        col = c
                        if aux_names_dict.get(c):
                            if aux_names_dict[c].get('comment'):
                                col = aux_names_dict[c].get('comment')
                            if aux_names_dict[c].get('label')\
                                and aux_names_dict[c].get('use_label') is True:
                                col = aux_names_dict[c].get('label')
                        new_names.append(col)
                    df.columns = new_names
                # EXPORT FILE
                _format = _conf.get('format') if _conf.get('format') else 'csv'
                name = _conf.get('name') if _conf.get('name') else 'export'               
                upload_path = self.conf.get('UPLOAD')
                temp = 'tmp'
                try:
                    os.stat(f'{os.getcwd()}/{upload_path}/{temp}')
                except Exception as _err:# pylint: disable=broad-exception-caught
                    os.mkdir(f'{os.getcwd()}/{upload_path}/{temp}')
                fname = f'{name}.{_format}'
                _path = f'{os.getcwd()}/{upload_path}/{temp}/{fname}'
                if _format in ['csv', '.csv']:
                    if _conf.get('compress') and _conf.get('compress_format'):
                        compress_format = _conf.get('compress_format')
                        fname = f'{name}.{_format}.{compress_format}'
                        _path = f'{os.getcwd()}/{upload_path}/{temp}/{fname}'
                    df.to_csv(_path, index = False)
                elif _format in ['xlsx', '.xlsx']:
                    df.to_excel(_path, index = False)
                return {'success': True, 'msg': self.i18n('success'), 'fname': f'{temp}/{fname}'}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # DUMPS FILE CONTENT TO JSON AND RETURN IT dump_file_2_object
    async def dump_file_2_object(self):
        '''dump file to an json'''
        try:
            fname = self.params['data'].get('file')
            is_tmp = self.params['data'].get('tmp')
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}'
            if is_tmp is True:
                _path = tempfile.gettempdir()
            #print(fname)
            basename, ext = os.path.splitext(fname)
            df = pd.DataFrame([])
            if ext in ['.csv']:
                df = pd.read_csv(f'{_path}/{fname}')
            elif ext in ['.json']:
                df = pd.read_json(f'{_path}/{fname}')
            elif ext in ['.xlsx', '.xls', '.xlsm']:
                df = pd.read_excel(f'{_path}/{fname}')
            elif ext in ['.xlsb']:
                df = pd.read_excel(f'{_path}/{fname}', engine = 'pyxlsb')
            else:
                return {'success': False, 'msg': self.i18n('extention-not-suported', ext = ext)}
            df = df.fillna('')
            df.replace(np.nan, '', regex = True, inplace = True)
            #print(df.shape, basename, ext, _path)
            data = df.to_dict(orient = 'records')
            return {'success': True, 'msg': self.i18n('success'), 'data': data, 'fname': fname}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # DUMPS THE QUERY TO THE DESIRE FORMAT AND RETURNS THE LINK FOR DOWNLOAD
    async def query(self):
        '''dump query'''
        try:
            try:
                engine = self.db.get_engine()
                # print('export/query', engine.url.drivername, engine.url)
                if engine.url.drivername in ['duckdb', 'duckdb_engine']:
                    self.params['only_return_str'] = True
            except Exception as err_:
                print(str(err_))
            _crud = Crud(self.conf, self.params, self.db, self.i18n)
            _query = await _crud.query()
            if not _query.get('success'):
                return _query
            else:
                _conf = self.params['data'].get('_conf')
                _format = _conf.get('format') if _conf.get('format') else 'csv'
                name = _conf.get('name', _conf.get('file', _conf.get('fname', 'export')))
                basename, ext = os.path.splitext(name)
                if not ext and _format:
                    fname = f'{name}.{_format}'
                elif not _format:
                    fname = f'{name}.csv'
                if self.params['only_return_str'] is True:
                    _sql_query = _query.get('sql', _query.get('query'))
                    # _qd = QueryDoc({})
                    self.params['data']['sql'] = _sql_query
                    self.params['data']['_conf'] = _conf
                    self.params['data']['tmp'] = True
                    self.params['data']['file'] = fname
                    self.params['data']['compress_format'] = _conf.get('compress_format')
                    _exp_class = ExportDucdb(self.conf, self.params, self.db, self.i18n)
                    return await _exp_class.export()
                _df = pd.DataFrame(_query.get('data'))             
                upload_path = self.conf.get('UPLOAD')
                temp = 'tmp'
                try:
                    os.stat(f'{os.getcwd()}/{upload_path}/{temp}')
                except FileExistsError as _err:
                    os.mkdir(f'{os.getcwd()}/{upload_path}/{temp}')
                _path = f'{os.getcwd()}/{upload_path}/{temp}/{fname}'
                if _format in ['csv', '.csv', 'parquet', '.parquet']:
                    if _conf.get('compress') and _conf.get('compress_format'):
                        compress_format = _conf.get('compress_format')
                        fname = f'{name}.{_format}.{compress_format}'
                        _path = f'{os.getcwd()}/{upload_path}/{temp}/{fname}'
                if _format in ['csv', '.csv']:
                    _other_csv_arg = {}
                    if _conf.get('csv'):
                        _other_csv_arg = _conf.get('csv')
                    _df.to_csv(_path, index = False, **_other_csv_arg)
                elif _format in ['xlsx', '.xlsx']:
                    _df.to_excel(_path, index = False)
                elif _format in ['parquet', '.parquet']:
                    _df.to_parquet(_path, index = False, engine = 'pyarrow')
                return {'success': True, 'msg': self.i18n('success'), 'fname': f'{temp}/{fname}'}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # DUMP TO HTML JINJA TEMPLATE
    async def dump_2_html(self):
        '''dump html jinja2 tmpl'''
        try:
            #html_str = self.template(env.from_string(self.params.get('htmlstr')),  data = self.params.get('data'))
            #print(self.params.get('htmlstr') )
            def handle_exception(value):
                try:
                    return value
                except Exception as _err:# pylint: disable=broad-exception-caught
                    return None
            _data = self.params.get('data', {})
            _env = Environment(loader = BaseLoader()).from_string(_data.get('htmlstr'))
            #_env.filters['handle_exception'] = handle_exception
            html_str = _env.render(**{'data': _data.get('data'), 'conf':  _data.get('_conf')})
            return {'success': True, 'msg': self.i18n('success'), 'html': html_str}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    async def dump_html_pdf(self):
        '''dump HTML to PDF depends on https://wkhtmltopdf.org/downloads.html'''
        try:
            _data = self.params.get('data', {})
            fname = _data.get('file')
            is_tmp = _data.get('tmp')
            _path = f'{os.getcwd()}/{self.conf.get("UPLOAD")}'
            if is_tmp is True:
                _path = tempfile.gettempdir()
            #html_str = self.template(env.from_string(self.params.get('htmlstr')),  data = self.params.get('data'))
            _env = Environment(loader = BaseLoader()).from_string(_data.get('htmlstr'))
            html_str = _env.render(**{'data': _data.get('data'), 'conf':  _data.get('_conf')})
            options = _data.get('options') if _data.get('options') else {}
            css = _data.get('css') if _data.get('css') else None
            pdfkit.from_string(html_str, f'{_path}/{fname}', options = options, css = css)
            return {'success': True, 'msg': self.i18n('success'), 'file': f'{fname}.pdf'}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
    # GENERIC EXPORT
    async def _get_new_db(self, _database):
        '''get db connection'''
        if isinstance(_database, dict):
            _patt = re.compile(r'@ENV\..+')
            for _key in _database:
                match_env = re.findall(_patt, str(_database[_key]))
                if len(match_env) > 0:
                    _env = re.sub(r'@ENV\.', '', str(match_env[0]))
                    try:
                        _database[_key] = os.environ.get(_env)
                    except Exception as _err:# pylint: disable=broad-exception-caught
                        pass
            patt = re.compile(r"\{.*?\}", re.IGNORECASE)
            matchs_odbc = re.findall(patt, _database.get('drivername'))
            if len(matchs_odbc) > 0:
                db_api = _database.get('odbc_dbapi')
                _prms = _database.get('odbc_url').format(**_database)
                #print(_prms)
                prms = urllib.parse.quote_plus(_prms)
                return create_engine(f"{db_api}+pyodbc:///?odbc_connect={prms}")
            else:
                # self.params['database'] = _database
                return DB(self.conf, self.params)
        return DB(self.conf, self.params)
    def find_edges(self, sheet):
        """FIND EDGES IN .XLSX"""
        row = sheet.max_row
        while row > 0:
            cells = sheet[row]
            if all([cell.value is None for cell in cells]):
                row -= 1
            else:
                break
        if row == 0:
            return 0, 0
        column = sheet.max_column
        while column > 0:
            cells = next(sheet.iter_cols(min_col=column, max_col=column, max_row=row))
            if all([cell.value is None for cell in cells]):
                column -= 1
            else:
                break
        return row, column
    def apply_format(self, _ws, formats):
        '''apply format'''
        for fmt in formats:
            if fmt.get('num_format') and fmt.get('range'):
                for cell in _ws[fmt.get('range')]:
                    if isinstance(cell, tuple):
                        for c in cell:
                            try:
                                c.number_format = fmt.get('num_format')
                            except Exception as _err:# pylint: disable=broad-exception-caught
                                print(str(_err))
                    else:
                        try:
                            cell.number_format = fmt.get('num_format')
                        except Exception as _err:# pylint: disable=broad-exception-caught
                            print(str(_err))
    def _rem_special_char(self, table, v):
        try:
            if not v:
                return v
            elif isinstance(v, str):
                return str(v).translate(table)
            else:
                return v
        except Exception as _err:
            return v
    async def export(self):
        '''GENERIC EXPORT'''
        try:            
            _data = self.params['data']
            _conf = _data.get('_conf', _data.get('conf', {}))
            _export = _data.get('export', [])
            _details = _data.get('data', [])
            fname = _data.get('file')
            template = _data.get('template')
            date_ref = _data.get('date_ref') if isinstance(_data.get('date_ref'), (datetime.datetime, datetime.date)) else parser.parse(_data.get('date_ref'))
            basename, ext = os.path.splitext(fname)
            # is_tmp = _data.get('tmp')
            _format = _conf.get('format') if _conf.get('format') else 'csv'
            if not ext and not _export.get('txt_fix_format_layout'):
                ext = _format
                fname = f'{fname}.{_format}'
            elif _export.get('txt_fix_format_layout'):
                ext = 'txt'
                _format = ext
                fname = f'{fname}.{_format}'
            else:
                _format = ext
            upload_path = self.conf.get('UPLOAD')
            tmp = 'tmp'
            try:
                os.stat(f'{os.getcwd()}/{upload_path}/{tmp}')
            except FileExistsError as _err:
                os.mkdir(f'{os.getcwd()}/{upload_path}/{tmp}')
            _path = f'{os.getcwd()}/{upload_path}/{tmp}/{fname}'
            _database = copy.deepcopy(self.conf.get('DATABASE'))
            _database['database'] = copy.deepcopy(self.params.get('database', self.params.get('db', None)))
            if _conf.get('params'):
                _database = copy.deepcopy(_conf.get('params'))
            _db = await self._get_new_db(_database)
            if isinstance(_db, sa.engine.base.Engine):
                engine = _db
            else:
                engine = _db.get_engine(_database)
            with engine.connect() as conn:
                try:
                    if engine.driver in ('pysqlite', 'sqlite'):
                        cache_size = self.conf.get('SQLITE_CACHE_SIZE', -2 * 1024 * 1024)
                        conn.execute(text(f'PRAGMA cache_size = {str(cache_size)}'))
                        busy_timeout = self.conf.get('SQLITE_BUSY_TIMEOUT', 60 * 1000) # 60s / 1m
                        conn.execute(text(f'PRAGMA busy_timeout = {busy_timeout}'))
                    _df = pd.DataFrame([])
                    if ext in ['csv', '.csv']:
                        if _conf.get('compress') and _conf.get('compress_format'):
                            compress_format = _conf.get('compress_format')
                            fname = f'{fname}.{compress_format}'
                            _path = f'{os.getcwd()}/{upload_path}/{tmp}/{fname}'
                        _other_csv_arg = {}
                        if _conf.get('csv'):
                            _other_csv_arg = _conf.get('csv')
                        _df = pd.read_sql(text(_details[0].get('sql_export_query')), con= conn)
                        _df.to_csv(_path, index = False, **_other_csv_arg)
                    elif ext in ['xlsx', '.xlsx', '.xlsm', 'xlsm']:
                        _tmpl_path = f'{os.getcwd()}/{upload_path}/{template}'
                        file_exists = os.path.exists(_tmpl_path)
                        if not file_exists: # DUMP EACH QUERY TO A SHEET
                            with pd.ExcelWriter(_path) as writer: # pylint: disable=abstract-class-instantiated
                                for _detail in _details:
                                    _df_aux = pd.read_sql(text(_detail.get('sql_export_query')), con = conn)
                                    _df_aux.to_excel(writer, sheet_name = _detail.get('dest_sheet_name'), index = False)
                                #writer._save()
                                writer.close()
                        else: # FILL TEMPLATE
                            try:
                                _wb = xl.Workbook()
                                _wb.remove(_wb[_wb.active])
                                if ext not in ['.xlsm', 'xlsm']:
                                    _wb = xl.load_workbook(filename = _tmpl_path, keep_links = False)
                                else:
                                    _wb = xl.load_workbook(filename = _tmpl_path, keep_links = False, keep_vba = True)
                                #xl_writer = pd.ExcelWriter(fname, engine = 'openpyxl') # pylint: disable=abstract-class-instantiated
                                #xl_writer.book = _wb
                                if _conf('type') in ['xlwriter'] and _conf('sheets'):
                                    _wb.save(fname)
                                    _wb.close()
                                    with pd.ExcelWriter(fname, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'overlay') as xl_writer:
                                        _conf_sheets = _conf('sheets', {})
                                        for _detail in _details:
                                            _sheet_name = _detail.get('dest_sheet_name')
                                            _args = {}
                                            _aux = {'startcol': 0, 'startrow': 2}
                                            for _key in ['index', 'header', 'startcol', 'startrow']:
                                                _args[_key] = _conf_sheets.get(_sheet_name, {}).get(_key, _aux.get(_key, False))
                                            formats = _conf_sheets.get(_sheet_name, {}).get('formats', None)
                                            _df_aux = pd.read_sql(text(_detail.get('sql_export_query')), con = conn)
                                            try:
                                                _df_aux.to_excel(xl_writer, sheet_name = _sheet_name, **_args)
                                                if formats:
                                                    try:
                                                        self.apply_format(xl_writer.sheets[_sheet_name], formats)
                                                    except Exception as _err2:# pylint: disable=broad-exception-caught
                                                        str(_err2) 
                                            except Exception as _err:# pylint: disable=broad-exception-caught
                                                str(_err)                                    
                                        try:
                                            _wb.save(fname)
                                            for ws_name in _wb.sheetnames:
                                                _ws = _wb[ws_name]
                                                if _ws._pivots: # pylint: disable=protected-access
                                                    print(ws_name, 'HAS PIVOTS:')
                                                    for pvt in _ws._pivots: # pylint: disable=protected-access
                                                        print('PIVOT:', pvt.name)
                                                        pvt.cache.refreshOnLoad = True
                                            _wb.close()
                                        except Exception as _err:# pylint: disable=broad-exception-caught
                                            *_, exc_tb = sys.exc_info()
                                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
                                elif _conf('type') in ['spreadsheet_form']:
                                    pass
                                else:
                                    for _detail in _details:
                                        _df_aux = pd.read_sql(text(_detail.get('sql_export_query')), con = conn)
                                        try:
                                            _wb.remove(_wb[_detail.get('dest_sheet_name')])
                                        except Exception as _err:# pylint: disable=broad-exception-caught
                                            str(_err)
                                        _ws = _wb.create_sheet(_detail.get('dest_sheet_name'))
                                        for _row in dataframe_to_rows(_df_aux, index = False, header = True):
                                            _ws.append(_row)
                                        row, column = self.find_edges(_ws)
                                        ref_cell = _ws.cell(row = row, column = column)
                                        tab = Table(displayName = _detail.get('dest_table_name'), ref = "A1:" + ref_cell.column_letter + str(ref_cell.row))
                                        style = TableStyleInfo(
                                            name = None,
                                            showFirstColumn = False,
                                            showLastColumn = False,
                                            showRowStripes = True,
                                            showColumnStripes = True
                                        )
                                        tab.tableStyleInfo = style
                                        _ws.add_table(tab)
                                        try:
                                            for ws_name in _wb.sheetnames:
                                                _ws = _wb[ws_name]
                                                if _ws._pivots:# pylint: disable=protected-access
                                                    print(ws_name, 'HAS PIVOTS:')
                                                    for pvt in _ws._pivots:# pylint: disable=protected-access
                                                        print('PIVOT:', pvt.name)
                                                        pvt.cache.refreshOnLoad = True
                                        except Exception as _err:# pylint: disable=broad-exception-caught
                                            *_, exc_tb = sys.exc_info()
                                            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
                                        _wb.save(fname)
                                        _wb.close()
                            except Exception as _err:# pylint: disable=broad-exception-caught
                                *_, exc_tb = sys.exc_info()
                                _fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                                print('DEBUG INF RECONC CREATION: ', str(_err), _fname, exc_tb.tb_lineno)
                                return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
                    elif ext in ['txt', '.txt'] and _export.get('txt_fix_format_layout'):
                        patt = re.compile(r'^.+\[.+\]$') # PATTERN TableName[FieldName]
                        patt2 = re.compile(r'\[.+\]$') # PATTERN FIELD [FieldName]
                        patt3 = re.compile(r'\[|\]') # PATTERN 2 MATCH [ or ]
                        df_layout = pd.DataFrame([])
                        is_table = False
                        txt = _export.get('txt_fix_format_layout')
                        matches = re.findall(patt, txt)
                        tbl = ''
                        fld = ''
                        if len(matches) > 0:
                            matches2 = re.findall(patt2, matches[0])
                            if len(matches2) > 0:
                                is_table = True
                                tbl = re.sub(patt2, '', matches[0])
                                fld = re.sub(patt3, '', matches2[0])
                        if is_table:
                            sql = f"""SELECT * FROM "{tbl}" """
                            df_layout = pd.read_sql(text(sql) , con = conn)
                            comp = df_layout[fld].values
                            #print(txt, tbl, fld,  comp)
                        else:                                      
                            comp = json.loads(f"[{_export.get('txt_fix_format_layout')}]")
                        txt = _export.get('txt_fix_format_header')
                        matches = re.findall(patt, txt)
                        tbl2 = ''
                        fld2 = ''
                        if len(matches) > 0:
                            matches2 = re.findall(patt2, matches[0])
                            if len(matches2) > 0:
                                is_table = True
                                tbl2 = re.sub(patt2, '', matches[0])
                                fld2 = re.sub(patt3, '', matches2[0])
                        if is_table:
                            if tbl2 != tbl:
                                sql = f"""SELECT * FROM "{tbl}" """
                                df_layout = pd.read_sql(text(sql) , con = conn)
                                headers = df_layout[fld].values
                            else:
                                headers = df_layout[fld].values
                            #print(txt, tbl, fld,  comp)
                        else:                                      
                            headers = json.loads(f"[{_export.get('txt_fix_format_header')}]")
                        _df = pd.read_sql(text(_details[0].get('sql_export_query')), con = conn)
                        _df.replace(np.nan, '', regex = True, inplace = True)
                        encod = 'utf-8'
                        fmts = ''
                        i = 0
                        for s in comp:
                            spc = ''
                            if i > 0:
                                spc = ''
                            align = ''
                            fmts += spc + '%' + align + str(s) + 's'
                            i += 1
                        try:
                            # pylint: disable=duplicate-key
                            normalizeChars = {
                                'Š' : 'S', 'š' : 's', 'Ð' : 'D', 'Ž' : 'Z', 'ž' : 'z', 'À' : 'A', 'Á' : 'A', 'Â' : 'A', 'Ã' : 'A', 'Ä' : 'A',
                                'Å' : 'A', 'Æ' : 'A', 'Ç' : 'C', 'È' : 'E', 'É' : 'E', 'Ê' : 'E', 'Ë' : 'E', 'Ì' : 'I', 'Í' : 'I', 'Î' : 'I',
                                'Ï' : 'I', 'Ñ' : 'N', 'Ń' : 'N', 'Ò' : 'O', 'Ó' : 'O', 'Ô' : 'O', 'Õ' : 'O', 'Ö' : 'O', 'Ø' : 'O', 'Ù' : 'U', 'Ú' : 'U',
                                'Û' : 'U', 'Ü' : 'U', 'Ý' : 'Y', 'Þ' : 'B', 'ß' : 'S', 'à' : 'a', 'á' : 'a', 'â' : 'a', 'ã' : 'a', 'ä' : 'a',
                                'å' : 'a', 'æ' : 'a', 'ç' : 'c', 'è' : 'e', 'é' : 'e', 'ê' : 'e', 'ë' : 'e', 'ì' : 'i', 'í' : 'i', 'î' : 'i',
                                'ï' : 'i', 'ð' : 'o', 'ñ' : 'n', 'ń' : 'n', 'ò' : 'o', 'ó' : 'o', 'ô' : 'o', 'õ' : 'o', 'ö' : 'o', 'ø' : 'o', 'ù' : 'u',
                                'ú' : 'u', 'û' : 'u', 'ü' : 'u', 'ý' : 'y', 'ý' : 'y', 'þ' : 'b', 'ÿ' : 'y', 'ƒ' : 'f',
                                'ă' : 'a', 'î' : 'i', 'â' : 'a', 'ș' : 's', 'ț' : 't', 'Ă' : 'A', 'Î' : 'I', 'Â' : 'A', 'Ș' : 'S', 'Ț' : 'T',
                                'º' : 'o', 'ª' : 'a', '�' : 'O', 'å': 'O'
                            }
                            def joinList(array):
                                join = ''
                                for item in array:
                                    join += str(item)
                                return join        
                            table = str.maketrans(joinList(normalizeChars.keys()), joinList(normalizeChars.values()))
                            for _f in _df.columns:
                                try:
                                    _df[_f] = np.vectorize(self._rem_special_char)(table, _df[_f])
                                except Exception as _err:
                                    pass
                        except Exception as _err:
                            pass
                        df2npArray = _df.values
                        np.savetxt(fname, df2npArray[:, 0:len(comp)], fmt = fmts, delimiter = '', encoding = encod)
                    return {'success': True, 'msg': self.i18n('success'), 'fname': f'{tmp}/{fname}'}
                except Exception as _err:# pylint: disable=broad-exception-caught
                    *_, exc_tb = sys.exc_info()
                    _fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print('DEBUG INF RECONC CREATION: ', str(_err), _fname, exc_tb.tb_lineno)
                    return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
        except Exception as _err:# pylint: disable=broad-exception-caught
            *_, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print('DEBUG INF: ', str(_err), fname, exc_tb.tb_lineno)
            return {'success': False, 'msg': self.i18n('unexpected-error', err = str(_err))}
